"""
Owner portal views for Atletas Performance Center.

All views are gated by @user_passes_test(is_owner), which allows access to
Django superusers, staff users, and members of the 'Owner' auth group.

Sections (separated by banner comments in the file):
  - Dashboard       — KPI tiles, revenue summary, recent activity
  - Packages        — CRUD for purchasable packages (add/edit/delete/restore/duplicate)
  - Session Types   — CRUD for session format catalogue + homepage display config
  - Coaches         — Add/edit/deactivate coaches; view schedule
  - Clients         — Browse clients; approve/reject coach & renter access
  - Players         — Browse players; view assessments
  - Bookings        — View and manage all bookings
  - Field Rental    — Manage rental slots, approve/reject tenant requests
  - Teams           — Team roster and booking overview
  - Finances        — Revenue reports, payment history, refunds
  - Credits         — Grant and track client credits
  - Discount Codes  — Create and monitor promo codes
  - Services        — Rental service catalogue
  - Waivers         — View signed waivers
  - Notifications   — Bulk email/push notification broadcasts
  - Contacts        — Pre-registration contact import list
  - Guide           — Owner how-to reference page

URL prefix: /owner-portal/  (all routes registered in atletasworld/urls.py)
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db.models import Count, Sum, Avg, Q, Case, When, Value, DecimalField
from datetime import timedelta

from coaches.models import Coach, ScheduleBlock, PlayerAssessment
from bookings.models import Booking, SessionType
from clients.models import Client, Player, ClientCredit, Package, ClientPackage, ClientWaiver, FieldRentalSlot, get_current_waiver
from reviews.models import Review
from django.contrib.auth.models import User
from django.core.mail import send_mass_mail, send_mail, EmailMessage
from email.mime.image import MIMEImage
from django.conf import settings
from django.views.decorators.http import require_POST


def is_owner(user):
    """Check if user is staff/superuser or in Owner group."""
    return user.is_staff or user.is_superuser or user.groups.filter(name='Owner').exists()


@login_required
@user_passes_test(is_owner)
def owner_dashboard(request):
    """Owner dashboard with overview across all entities."""
    today = timezone.localdate()
    month_start = today.replace(day=1)
    year_start  = today.replace(month=1, day=1)

    # Aware datetime boundaries for DateTimeField comparisons (avoids naive-datetime warnings)
    def _dt(d):
        from datetime import datetime as _datetime
        return timezone.make_aware(_datetime(d.year, d.month, d.day))

    month_start_dt     = _dt(month_start)
    year_start_dt      = _dt(year_start)

    # ── Core counts ────────────────────────────────────────────────────────────
    total_coaches  = Coach.objects.filter(is_active=True).count()
    _client_qs = Client.objects.filter(user__is_staff=False, user__is_superuser=False
                   ).exclude(user__groups__name__in=['Owner', 'Coach'])
    total_clients  = _client_qs.count()
    total_players  = Player.objects.filter(is_active=True).count()
    pending_approvals = _client_qs.filter(approval_status='pending').count()

    # ── Today ──────────────────────────────────────────────────────────────────
    todays_bookings      = Booking.objects.filter(scheduled_date=today).count()
    total_sessions_today = ScheduleBlock.objects.filter(date=today).count()
    pending_bookings     = Booking.objects.filter(status__in=['pending', 'confirmed'],
                                                   scheduled_date__gte=today).count()

    # ── Financial ──────────────────────────────────────────────────────────────
    last_month_start    = (month_start - timedelta(days=1)).replace(day=1)
    last_month_start_dt = _dt(last_month_start)
    today_dt            = _dt(today + timedelta(days=1))  # exclusive upper bound (end of today)

    # 1 query: all booking drop-in revenue periods via conditional Sum
    # scheduled_date is DateField so date comparisons are correct here
    booking_rev = Booking.objects.filter(payment_status='paid').aggregate(
        this_month=Sum(Case(When(scheduled_date__gte=month_start, scheduled_date__lte=today,   then='amount_paid'), default=Value(0), output_field=DecimalField())),
        ytd=       Sum(Case(When(scheduled_date__gte=year_start,  scheduled_date__lte=today,   then='amount_paid'), default=Value(0), output_field=DecimalField())),
        last_month=Sum(Case(When(scheduled_date__gte=last_month_start, scheduled_date__lt=month_start, then='amount_paid'), default=Value(0), output_field=DecimalField())),
    )
    # 1 query: package revenue = actual amount charged (Payment.amount, status=succeeded)
    from payments.models import Payment as _Payment
    pkg_rev = _Payment.objects.filter(status='succeeded').aggregate(
        this_month=Sum(Case(When(created_at__gte=month_start_dt, created_at__lt=today_dt,        then='amount'), default=Value(0), output_field=DecimalField())),
        ytd=       Sum(Case(When(created_at__gte=year_start_dt,  created_at__lt=today_dt,        then='amount'), default=Value(0), output_field=DecimalField())),
        last_month=Sum(Case(When(created_at__gte=last_month_start_dt, created_at__lt=month_start_dt, then='amount'), default=Value(0), output_field=DecimalField())),
    )
    # 1 query: all rental revenue periods (approved_at is DateTimeField — use aware datetimes)
    rental_rev = FieldRentalSlot.objects.filter(payment_status='paid').aggregate(
        this_month=Sum(Case(When(approved_at__gte=month_start_dt, approved_at__lt=today_dt,        then='amount_paid'), default=Value(0), output_field=DecimalField())),
        ytd=       Sum(Case(When(approved_at__gte=year_start_dt,  approved_at__lt=today_dt,        then='amount_paid'), default=Value(0), output_field=DecimalField())),
        last_month=Sum(Case(When(approved_at__gte=last_month_start_dt, approved_at__lt=month_start_dt, then='amount_paid'), default=Value(0), output_field=DecimalField())),
    )
    revenue_this_month = (booking_rev['this_month'] or 0) + (pkg_rev['this_month'] or 0) + (rental_rev['this_month'] or 0)
    revenue_ytd        = (booking_rev['ytd']        or 0) + (pkg_rev['ytd']        or 0) + (rental_rev['ytd']        or 0)
    revenue_last_month = (booking_rev['last_month'] or 0) + (pkg_rev['last_month'] or 0) + (rental_rev['last_month'] or 0)
    pending_payments_qs = Booking.objects.filter(
        payment_status='pending', status__in=['pending', 'confirmed']
    ).select_related('client__user', 'player', 'session_type', 'coach__user').order_by('scheduled_date')
    pending_payments = pending_payments_qs.aggregate(t=Sum('amount_paid'))['t'] or 0
    rental_revenue_month = FieldRentalSlot.objects.filter(
        status='booked',
        date__gte=month_start, date__lte=today
    ).aggregate(t=Sum('service__price'))['t'] or 0

    # Recent paid transactions — drop-ins + package purchases
    _raw_bookings = Booking.objects.filter(
        payment_status='paid', amount_paid__gt=0
    ).select_related('client__user', 'player', 'session_type').order_by('-updated_at')[:8]
    _raw_packages = ClientPackage.objects.exclude(
        status='cancelled'
    ).select_related('client__user', 'package', 'player').order_by('-purchase_date')[:8]
    _transactions = []
    for bk in _raw_bookings:
        _transactions.append({
            'name': (f"{bk.player.first_name} {bk.player.last_name}".strip()
                     if bk.player else bk.client.user.get_full_name()),
            'label': bk.session_type.name if bk.session_type else 'Session',
            'amount': bk.amount_paid,
            'date': bk.updated_at.date(),
            'type': 'dropin',
        })
    for cp in _raw_packages:
        _transactions.append({
            'name': (f"{cp.player.first_name} {cp.player.last_name}".strip()
                     if cp.player else cp.client.user.get_full_name()),
            'label': cp.package.name,
            'amount': cp.package.price,
            'date': cp.purchase_date.date(),
            'type': 'package',
        })
    _transactions.sort(key=lambda x: x['date'], reverse=True)
    recent_transactions = _transactions[:8]

    # ── Coaches ────────────────────────────────────────────────────────────────
    coaches = Coach.objects.filter(is_active=True).annotate(
        sessions_today=Count('schedule_blocks', filter=Q(schedule_blocks__date=today)),
        upcoming=Count('bookings', filter=Q(bookings__scheduled_date__gte=today,
                                            bookings__status__in=['pending','confirmed'])),
        total_players=Count('bookings__player', distinct=True)
    ).order_by('-sessions_today')[:10]

    todays_schedule = ScheduleBlock.objects.filter(
        date=today
    ).select_related('coach__user').order_by('start_time')[:20]

    recent_bookings = Booking.objects.select_related(
        'client__user', 'player', 'coach__user', 'session_type'
    ).order_by('-created_at')[:10]

    players_pending_assessment = Booking.objects.filter(
        status='completed',
        scheduled_date__gte=today - timedelta(days=14)
    ).exclude(assessments__isnull=False).select_related('player', 'coach__user')[:10]

    # ── Rentals ────────────────────────────────────────────────────────────────
    rentals_pending  = FieldRentalSlot.objects.filter(status='pending_approval').count()
    rentals_upcoming = FieldRentalSlot.objects.filter(
        status='booked', date__gte=today
    ).count()
    rentals_today = FieldRentalSlot.objects.filter(
        date=today, status__in=['booked','pending_approval']
    ).select_related('service').order_by('start_time')[:5]

    # ── Waivers ────────────────────────────────────────────────────────────────
    from clients.models import ClientWaiver
    current_year = today.year
    waiver_signed_count = ClientWaiver.objects.filter(
        valid_year=current_year,
        waiver_version=ClientWaiver.WAIVER_VERSION,
    ).values('client_id').distinct().count()
    waiver_unsigned_count = _client_qs.filter(
        user__groups__name='Client'
    ).exclude(
        waivers__valid_year=current_year,
        waivers__waiver_version=ClientWaiver.WAIVER_VERSION,
    ).count()

    # ── Active Packages ────────────────────────────────────────────────────────
    active_packages_count  = ClientPackage.objects.filter(
        status='active', expiry_date__gte=today
    ).count()
    expiring_soon_packages = ClientPackage.objects.filter(
        status='active',
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=7)
    ).select_related('client__user', 'package').order_by('expiry_date')[:8]
    packages_exhausted = ClientPackage.objects.filter(
        status='exhausted', expiry_date__gte=today
    ).count()

    # ── Stripe ─────────────────────────────────────────────────────────────────
    from payments.models import Payment
    stripe_confirmed = Payment.objects.filter(status='succeeded').aggregate(
        t=Sum('amount'))['t'] or 0
    stripe_count = Payment.objects.filter(status='succeeded').count()
    stripe_live = bool(settings.STRIPE_SECRET_KEY and
                       settings.STRIPE_SECRET_KEY.startswith(('sk_live', 'rk_live')))

    # ── Contacts ───────────────────────────────────────────────────────────────
    from clients.models import ContactParent
    contacts_unregistered = ContactParent.objects.filter(client__isnull=True).count()

    context = {
        'today': today,
        # Core counts
        'total_coaches': total_coaches,
        'total_clients': total_clients,
        'total_players': total_players,
        'pending_approvals': pending_approvals,
        # Today
        'todays_bookings': todays_bookings,
        'total_sessions_today': total_sessions_today,
        'pending_bookings': pending_bookings,
        # Financial
        'revenue_this_month': revenue_this_month,
        'revenue_ytd': revenue_ytd,
        'revenue_last_month': revenue_last_month,
        'pending_payments': pending_payments,
        'pending_payments_list': pending_payments_qs[:20],
        'rental_revenue_month': rental_revenue_month,
        'recent_transactions': recent_transactions,
        # Rentals
        'rentals_pending': rentals_pending,
        'rentals_upcoming': rentals_upcoming,
        'rentals_today': rentals_today,
        # Waivers
        'waiver_signed_count': waiver_signed_count,
        'waiver_unsigned_count': waiver_unsigned_count,
        'current_year': current_year,
        # Packages
        'active_packages_count': active_packages_count,
        'expiring_soon_packages': expiring_soon_packages,
        'packages_exhausted': packages_exhausted,
        # Stripe
        'stripe_live': stripe_live,
        'stripe_count': stripe_count,
        'stripe_confirmed': stripe_confirmed,
        # Contacts
        'contacts_unregistered': contacts_unregistered,
        # Lists
        'coaches': coaches,
        'todays_schedule': todays_schedule,
        'recent_bookings': recent_bookings,
        'players_pending_assessment': players_pending_assessment,
    }
    return render(request, 'owner/dashboard.html', context)


@login_required
@user_passes_test(is_owner)
def owner_notifications(request):
    """Owner notification center - send emails to different groups."""
    from clients.models import Package, ClientPackage, ContactParent, EmailBroadcast
    # Get counts for each recipient group
    all_clients = Client.objects.select_related('user').filter(user__email__isnull=False).exclude(user__email='')
    all_coaches = Coach.objects.select_related('user').filter(is_active=True, user__email__isnull=False).exclude(user__email='')
    all_users = User.objects.filter(is_active=True, email__isnull=False).exclude(email='')

    today = timezone.localdate()

    # Clients with bookings in last 30 days
    active_client_ids = Booking.objects.filter(
        scheduled_date__gte=today - timedelta(days=30)
    ).values_list('client_id', flat=True).distinct()
    active_clients = Client.objects.filter(id__in=active_client_ids).select_related('user')

    # Clients with bookings this week
    weekly_client_ids = Booking.objects.filter(
        scheduled_date__gte=today,
        scheduled_date__lte=today + timedelta(days=7)
    ).values_list('client_id', flat=True).distinct()
    clients_with_bookings_this_week = Client.objects.filter(id__in=weekly_client_ids).select_related('user')

    # Clients with any active package
    packaged_client_ids = ClientPackage.objects.filter(
        status='active',
        expiry_date__gte=today,
    ).values_list('client_id', flat=True).distinct()
    packaged_clients_count = Client.objects.filter(
        id__in=packaged_client_ids,
        user__email__isnull=False,
    ).exclude(user__email='').count()

    # Active packages with their active client counts (for per-package targeting)
    active_packages = Package.objects.filter(is_active=True).order_by('package_type', 'name')
    packages_with_counts = []
    for pkg in active_packages:
        count = ClientPackage.objects.filter(
            package=pkg,
            status='active',
            expiry_date__gte=today,
        ).values('client_id').distinct().count()
        packages_with_counts.append((pkg, count))

    # Contact list counts
    all_contacts     = ContactParent.objects.exclude(email='').order_by('last_name', 'first_name', 'email')
    unregistered_contacts = all_contacts.filter(client__isnull=True)
    contact_sources  = ContactParent.SOURCE_CHOICES

    # Per-source counts
    from django.db.models import Count as DjCount
    contacts_by_source = {
        row['source']: row['n']
        for row in ContactParent.objects.exclude(email='').values('source').annotate(n=DjCount('id'))
    }

    context = {
        'all_clients_count': all_clients.count(),
        'all_coaches_count': all_coaches.count(),
        'all_users_count': all_users.count(),
        'active_clients_count': active_clients.count(),
        'clients_with_bookings_this_week_count': clients_with_bookings_this_week.count(),
        'packaged_clients_count': packaged_clients_count,
        'packages_with_counts': packages_with_counts,
        'all_clients': all_clients,
        'all_coaches': all_coaches,
        # contact list
        'all_contacts': all_contacts,
        'all_contacts_count': all_contacts.count(),
        'unregistered_contacts_count': unregistered_contacts.count(),
        'contact_sources': contact_sources,
        'contacts_by_source': contacts_by_source,
        'recent_broadcasts': EmailBroadcast.objects.order_by('-created_at')[:10],
    }
    return render(request, 'owner/notifications.html', context)


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_send_notification(request):
    """Send notifications to selected recipients with optional attachments and images."""
    import logging
    logger = logging.getLogger(__name__)

    recipient_group = request.POST.get('recipient_group', '')
    subject = request.POST.get('subject', '').strip().replace('\n', '').replace('\r', '')
    message = request.POST.get('message', '').strip()
    individual_emails = request.POST.getlist('individual_emails')
    send_as_html = request.POST.get('send_as_html') == 'on'

    # Handle file uploads — filter out zero-byte entries (mobile browsers often submit empty file fields)
    attachments = [f for f in request.FILES.getlist('attachments') if f.size > 0]
    inline_image = request.FILES.get('inline_image')
    if inline_image and inline_image.size == 0:
        inline_image = None

    if not subject or not message:
        messages.error(request, 'Please provide both subject and message.')
        return redirect('owner_notifications')

    from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@atletasperformancecenter.com')

    try:
        if recipient_group == 'individual' and not individual_emails:
            messages.error(request, 'No recipients specified for individual send.')
            return redirect('owner_notifications')

        # Save uploaded files to disk so Celery can read them after the HTTP request ends.
        import uuid, os
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'email_attachments')
        os.makedirs(upload_dir, exist_ok=True)

        saved_attachments = []
        for att in attachments:
            safe_name = f"{uuid.uuid4().hex}_{att.name}"
            save_path = os.path.join(upload_dir, safe_name)
            with open(save_path, 'wb') as fh:
                for chunk in att.chunks():
                    fh.write(chunk)
            saved_attachments.append({'path': save_path, 'name': att.name, 'content_type': att.content_type})

        saved_inline_image = None
        if inline_image:
            safe_name = f"{uuid.uuid4().hex}_{inline_image.name}"
            save_path = os.path.join(upload_dir, safe_name)
            with open(save_path, 'wb') as fh:
                for chunk in inline_image.chunks():
                    fh.write(chunk)
            saved_inline_image = {'path': save_path, 'name': inline_image.name, 'content_type': inline_image.content_type}

        # Always dispatch to Celery — never block the HTTP request on email sends.
        from clients.models import EmailBroadcast
        from clients.tasks import send_bulk_email_task, run_task
        broadcast = EmailBroadcast.objects.create(
            recipient_group=recipient_group,
            subject=subject,
            sent_by=request.user,
        )
        run_task(send_bulk_email_task,
                 broadcast_id=broadcast.id,
                 recipient_group=recipient_group,
                 subject=subject,
                 message=message,
                 from_email=from_email,
                 send_as_html=send_as_html,
                 extra_params={
                     'package_id': request.POST.get('package_id', ''),
                     'contact_source': request.POST.get('contact_source', ''),
                     'individual_emails': list(individual_emails),
                     'attachments': saved_attachments,
                     'inline_image': saved_inline_image,
                 })
        messages.success(request,
            'Email queued for sending. Check "Recent Sends" below for delivery results.')

    except Exception as e:
        logger.error(f'owner_send_notification error: {e}', exc_info=True)
        messages.error(request, f'Error preparing email: {str(e)}')

    return redirect('owner_notifications')


def _resolve_recipient_emails(recipient_group, package_id='', contact_source='', individual_emails=None):
    """Resolve a recipient group name to a set of email addresses."""
    from bookings.models import Booking
    recipients = set()
    today = timezone.localdate()

    if recipient_group == 'all_clients':
        emails = Client.objects.filter(
            user__email__isnull=False
        ).exclude(user__email='').values_list('user__email', flat=True)
        recipients.update(emails)

    elif recipient_group == 'all_coaches':
        emails = Coach.objects.filter(
            is_active=True,
            user__email__isnull=False
        ).exclude(user__email='').values_list('user__email', flat=True)
        recipients.update(emails)

    elif recipient_group == 'everyone':
        emails = User.objects.filter(
            is_active=True,
            email__isnull=False
        ).exclude(email='').values_list('email', flat=True)
        recipients.update(emails)

    elif recipient_group == 'active_clients':
        active_client_ids = Booking.objects.filter(
            scheduled_date__gte=today - timedelta(days=30)
        ).values_list('client_id', flat=True).distinct()
        emails = Client.objects.filter(
            id__in=active_client_ids,
            user__email__isnull=False
        ).exclude(user__email='').values_list('user__email', flat=True)
        recipients.update(emails)

    elif recipient_group == 'clients_this_week':
        weekly_client_ids = Booking.objects.filter(
            scheduled_date__gte=today,
            scheduled_date__lte=today + timedelta(days=7)
        ).values_list('client_id', flat=True).distinct()
        emails = Client.objects.filter(
            id__in=weekly_client_ids,
            user__email__isnull=False
        ).exclude(user__email='').values_list('user__email', flat=True)
        recipients.update(emails)

    elif recipient_group == 'packaged_clients':
        from clients.models import ClientPackage
        packaged_ids = ClientPackage.objects.filter(
            status='active',
            expiry_date__gte=today,
        ).values_list('client_id', flat=True).distinct()
        emails = Client.objects.filter(
            id__in=packaged_ids,
            user__email__isnull=False,
        ).exclude(user__email='').values_list('user__email', flat=True)
        recipients.update(emails)

    elif recipient_group == 'package_specific':
        from clients.models import ClientPackage
        if package_id:
            packaged_ids = ClientPackage.objects.filter(
                package_id=package_id,
                status='active',
                expiry_date__gte=today,
            ).values_list('client_id', flat=True).distinct()
            emails = Client.objects.filter(
                id__in=packaged_ids,
                user__email__isnull=False,
            ).exclude(user__email='').values_list('user__email', flat=True)
            recipients.update(emails)

    elif recipient_group == 'contacts_all':
        from clients.models import ContactParent
        emails = ContactParent.objects.exclude(email='').values_list('email', flat=True)
        recipients.update(emails)

    elif recipient_group == 'contacts_unregistered':
        from clients.models import ContactParent
        emails = ContactParent.objects.filter(
            client__isnull=True
        ).exclude(email='').values_list('email', flat=True)
        recipients.update(emails)

    elif recipient_group == 'contacts_by_source':
        from clients.models import ContactParent
        if contact_source:
            emails = ContactParent.objects.filter(
                source=contact_source
            ).exclude(email='').values_list('email', flat=True)
            recipients.update(emails)

    elif recipient_group == 'individual':
        if individual_emails:
            recipients.update(individual_emails)

    return recipients


def _build_html_email(html_message, site_url):
    """Return branded APC HTML email string."""
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif; line-height: 1.6; color: #333333; background-color: #f5f5f5; margin: 0; padding: 0; }}
    .email-wrapper {{ max-width: 600px; margin: 0 auto; background-color: #ffffff; }}
    .email-header {{ background: linear-gradient(135deg, #1a1a1a 0%, #2c3e50 100%); padding: 30px; text-align: center; }}
    .email-header img {{ max-height: 60px; width: auto; }}
    .email-header h1 {{ color: #ffffff; margin: 12px 0 0 0; font-size: 22px; font-weight: 600; letter-spacing: 0.5px; }}
    .email-body {{ padding: 40px 30px; }}
    .email-body p {{ margin: 0 0 15px 0; color: #555555; }}
    .divider {{ border: none; border-top: 1px solid #eeeeee; margin: 30px 0; }}
    .signature {{ font-size: 14px; color: #444444; }}
    .signature strong {{ color: #1a1a1a; font-size: 15px; }}
    .signature .title {{ color: #888888; font-size: 13px; margin: 2px 0; }}
    .signature .contact {{ color: #888888; font-size: 13px; margin: 2px 0; }}
    .signature .contact a {{ color: #1a1a1a; text-decoration: none; }}
    .signature-bar {{ width: 40px; height: 3px; background-color: #D7FF00; margin: 10px 0; }}
    .email-footer {{ background-color: #1a1a1a; padding: 25px 30px; text-align: center; }}
    .email-footer p {{ color: #888888; font-size: 12px; margin: 4px 0; }}
    .email-footer a {{ color: #D7FF00; text-decoration: none; }}
    .footer-logo {{ color: #ffffff; font-size: 15px; font-weight: 700; letter-spacing: 1px; margin-bottom: 8px; }}
    @media only screen and (max-width: 600px) {{
        .email-body {{ padding: 25px 20px; }}
        .email-header {{ padding: 20px; }}
    }}
</style>
</head>
<body>
<div class="email-wrapper">
    <div class="email-header">
        <img src="{site_url}/static/img/apc-logo-yellow.png" alt="Atletas Performance Center" onerror="this.style.display='none'">
        <h1>Atletas Performance Center</h1>
    </div>
    <div class="email-body">
        {html_message}
        <hr class="divider">
        <div class="signature">
            <div class="signature-bar"></div>
            <strong>Atletas Performance Center</strong><br>
            <div class="title">Professional Soccer Training</div>
            <div class="contact">📧 <a href="mailto:info@atletasperformancecenter.com">info@atletasperformancecenter.com</a></div>
            <div class="contact">🌐 <a href="{site_url}">{site_url.replace("https://", "")}</a></div>
        </div>
    </div>
    <div class="email-footer">
        <div class="footer-logo">APC</div>
        <p>
            <a href="https://www.instagram.com/atletasworld/" target="_blank">Instagram</a> &nbsp;|&nbsp;
            <a href="https://www.facebook.com/atletasworld/" target="_blank">Facebook</a>
        </p>
        <p style="margin-top: 12px; font-size: 11px; color: #555555;">
            &copy; 2026 Atletas Performance Center. All rights reserved.
        </p>
    </div>
</div>
</body>
</html>'''


# ============================================================================
# PACKAGE MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_packages(request):
    """List all packages with management options."""
    from clients.models import Package, ClientPackage
    from django.db.models import Count

    base_qs = Package.objects.annotate(
        active_purchases=Count('clientpackage', filter=Q(clientpackage__status='active')),
        total_purchases=Count('clientpackage')
    )
    packages         = base_qs.filter(is_active=True).order_by('price')
    archived_packages= base_qs.filter(is_active=False).order_by('price')

    context = {
        'packages': packages,
        'archived_packages': archived_packages,
    }
    return render(request, 'owner/packages.html', context)


@login_required
@user_passes_test(is_owner)
def owner_package_add(request):
    """Add a new package."""
    from clients.models import Package

    if request.method == 'POST':
        try:
            package = Package.objects.create(
                name=request.POST.get('name'),
                package_type=request.POST.get('package_type'),
                description=request.POST.get('description', ''),
                price=request.POST.get('price'),
                sessions_included=request.POST.get('sessions_included', 0),
                validity_weeks=request.POST.get('validity_weeks', 4),
                is_active=request.POST.get('is_active') == 'on',
                is_purchasable=request.POST.get('is_purchasable') == 'on',
                is_special=request.POST.get('is_special') == 'on',
                max_participants=request.POST.get('max_participants', 0),
                age_group=request.POST.get('age_group', ''),
                event_start_date=request.POST.get('event_start_date') or None,
                event_start_time=request.POST.get('event_start_time') or None,
                event_end_date=request.POST.get('event_end_date') or None,
                event_end_time=request.POST.get('event_end_time') or None,
                event_location=request.POST.get('event_location', ''),
            )
            messages.success(request, f'Package "{package.name}" created successfully!')
            return redirect('owner_packages')
        except Exception as e:
            messages.error(request, f'Error creating package: {str(e)}')

    context = {
        'package_types': Package.PACKAGE_TYPE_CHOICES,
    }
    return render(request, 'owner/package_form.html', context)


@login_required
@user_passes_test(is_owner)
def owner_package_edit(request, pk):
    """Edit an existing package."""
    from clients.models import Package
    from django.shortcuts import get_object_or_404

    package = get_object_or_404(Package, pk=pk)

    if request.method == 'POST':
        try:
            package.name = request.POST.get('name')
            package.package_type = request.POST.get('package_type')
            package.description = request.POST.get('description', '')
            package.price = request.POST.get('price')
            package.sessions_included = request.POST.get('sessions_included', 0)
            package.validity_weeks = request.POST.get('validity_weeks', 4)
            package.is_active      = request.POST.get('is_active') == 'on'
            package.is_purchasable = request.POST.get('is_purchasable') == 'on'
            package.is_special     = request.POST.get('is_special') == 'on'
            package.max_participants = request.POST.get('max_participants', 0)
            package.age_group = request.POST.get('age_group', '')
            package.event_start_date = request.POST.get('event_start_date') or None
            package.event_start_time = request.POST.get('event_start_time') or None
            package.event_end_date = request.POST.get('event_end_date') or None
            package.event_end_time = request.POST.get('event_end_time') or None
            package.event_location = request.POST.get('event_location', '')
            package.save()
            messages.success(request, f'Package "{package.name}" updated successfully!')
            return redirect('owner_packages')
        except Exception as e:
            messages.error(request, f'Error updating package: {str(e)}')

    context = {
        'package': package,
        'package_types': Package.PACKAGE_TYPE_CHOICES,
        'editing': True,
    }
    return render(request, 'owner/package_form.html', context)


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_package_delete(request, pk):
    """Archive a package (soft delete — sets is_active=False)."""
    from clients.models import Package
    from django.shortcuts import get_object_or_404

    package = get_object_or_404(Package, pk=pk)
    package.is_active = False
    package.save()
    messages.success(request, f'Package "{package.name}" archived.')
    return redirect('owner_packages')


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_package_restore(request, pk):
    """Restore an archived package."""
    from clients.models import Package
    from django.shortcuts import get_object_or_404

    package = get_object_or_404(Package, pk=pk)
    package.is_active = True
    package.save()
    messages.success(request, f'Package "{package.name}" restored.')
    return redirect('owner_packages')


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_package_hard_delete(request, pk):
    """Permanently delete a package (only if no active client purchases)."""
    from clients.models import Package, ClientPackage
    from django.shortcuts import get_object_or_404

    package = get_object_or_404(Package, pk=pk)
    active_count = ClientPackage.objects.filter(package=package, status='active').count()
    if active_count > 0:
        messages.error(request, f'Cannot delete "{package.name}" — {active_count} active purchase(s) exist. Archive it instead.')
    else:
        name = package.name
        package.delete()
        messages.success(request, f'Package "{name}" permanently deleted.')
    return redirect('owner_packages')


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_package_duplicate(request, pk):
    """Duplicate a package."""
    from clients.models import Package
    from django.shortcuts import get_object_or_404

    orig = get_object_or_404(Package, pk=pk)
    copy = Package.objects.create(
        name=f'{orig.name} (Copy)',
        package_type=orig.package_type,
        description=orig.description,
        price=orig.price,
        sessions_included=orig.sessions_included,
        validity_weeks=orig.validity_weeks,
        is_active=False,  # start archived so owner can review before publishing
        is_special=orig.is_special,
        age_group=orig.age_group,
        max_participants=orig.max_participants,
    )
    messages.success(request, f'Package duplicated as "{copy.name}". Review and activate when ready.')
    return redirect('owner_package_edit', pk=copy.pk)


# ── Session Type actions ──────────────────────────────────────────────────────

@login_required
@user_passes_test(is_owner)
@require_POST
def owner_session_type_hard_delete(request, pk):
    """Permanently delete a session type (only if no bookings reference it)."""
    from bookings.models import SessionType, Booking
    from django.shortcuts import get_object_or_404

    st = get_object_or_404(SessionType, pk=pk)
    booking_count = Booking.objects.filter(session_type=st).count()
    if booking_count > 0:
        messages.error(request, f'Cannot delete "{st.name}" — {booking_count} booking(s) reference it. Archive it instead.')
    else:
        name = st.name
        st.delete()
        messages.success(request, f'Session type "{name}" permanently deleted.')
    return redirect('owner_session_types')


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_session_type_duplicate(request, pk):
    """Duplicate a session type."""
    from bookings.models import SessionType
    from django.shortcuts import get_object_or_404

    orig = get_object_or_404(SessionType, pk=pk)
    copy = SessionType.objects.create(
        name=f'{orig.name} (Copy)',
        description=orig.description,
        session_format=orig.session_format,
        duration_minutes=orig.duration_minutes,
        price=orig.price,
        drop_in_price=orig.drop_in_price,
        max_participants=orig.max_participants,
        color=orig.color,
        is_active=False,  # start archived
        requires_package=orig.requires_package,
        show_as_event=False,
        show_as_program=False,
        location=orig.location,
        age_group=orig.age_group,
        days_of_week=orig.days_of_week,
        start_times=orig.start_times,
        weekend_start_times=orig.weekend_start_times,
    )
    messages.success(request, f'Session type duplicated as "{copy.name}". Review and activate when ready.')
    return redirect('owner_session_type_edit', pk=copy.pk)


# ============================================================================
# COACH MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_coaches(request):
    """List all coaches with management options."""
    from django.contrib.auth.models import Group

    today = timezone.localdate()
    active_client_q = Q(bookings__client__approval_status__in=['approved', 'not_required'])
    coaches = Coach.objects.annotate(
        today_sessions=Count('bookings', filter=active_client_q & Q(
            bookings__scheduled_date=today,
            bookings__status__in=['pending', 'confirmed'],
        )),
        upcoming_sessions=Count('bookings', filter=active_client_q & Q(
            bookings__scheduled_date__gt=today,
            bookings__status__in=['pending', 'confirmed'],
        )),
        total_bookings=Count('bookings'),
        total_players=Count('bookings__player', distinct=True)
    ).order_by('-is_active', 'user__first_name')

    context = {
        'coaches': coaches,
    }
    return render(request, 'owner/coaches.html', context)


@login_required
@user_passes_test(is_owner)
def owner_coach_add(request):
    """Add a new coach."""
    from django.contrib.auth.models import Group

    if request.method == 'POST':
        try:
            # Create user
            username = request.POST.get('username')
            email = request.POST.get('email')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            password = request.POST.get('password')

            # Check if user exists
            if User.objects.filter(username=username).exists():
                messages.error(request, f'Username "{username}" already exists.')
                return redirect('owner_coach_add')
            if User.objects.filter(email=email).exists():
                messages.error(request, f'Email "{email}" already exists.')
                return redirect('owner_coach_add')

            user = User.objects.create_user(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                password=password
            )

            # Add to Coach group
            coach_group, _ = Group.objects.get_or_create(name='Coach')
            user.groups.add(coach_group)

            # Create coach profile
            coach = Coach.objects.create(
                user=user,
                slug=request.POST.get('slug', username),
                tagline=request.POST.get('tagline', '')[:200],
                bio=request.POST.get('bio', ''),
                full_bio=request.POST.get('full_bio', ''),
                specializations=request.POST.get('specializations', ''),
                certifications=request.POST.get('certifications', ''),
                experience_years=int(request.POST.get('experience_years', 0) or 0),
                coaching_philosophy=request.POST.get('coaching_philosophy', ''),
                achievements=request.POST.get('achievements', ''),
                instagram_url=request.POST.get('instagram_url', ''),
                facebook_url=request.POST.get('facebook_url', ''),
                twitter_url=request.POST.get('twitter_url', ''),
                linkedin_url=request.POST.get('linkedin_url', ''),
                hourly_rate=request.POST.get('hourly_rate', 0),
                is_active=request.POST.get('is_active') == 'on',
                profile_enabled=request.POST.get('profile_enabled') == 'on',
            )
            if 'photo' in request.FILES:
                coach.photo = request.FILES['photo']
                coach.save(update_fields=['photo'])

            messages.success(request, f'Coach "{first_name} {last_name}" created successfully!')
            return redirect('owner_coaches')
        except Exception as e:
            messages.error(request, f'Error creating coach: {str(e)}')

    return render(request, 'owner/coach_form.html', {'editing': False})


@login_required
@user_passes_test(is_owner)
def owner_coach_edit(request, pk):
    """Edit an existing coach."""
    from django.shortcuts import get_object_or_404

    coach = get_object_or_404(Coach, pk=pk)
    today = timezone.localdate()

    # Check for outstanding activities — 2 queries instead of 3
    booking_counts = Booking.objects.filter(coach=coach).aggregate(
        upcoming=Count('id', filter=Q(scheduled_date__gte=today, status__in=['pending', 'confirmed'])),
        pending_assess=Count('id', filter=Q(status='completed', scheduled_date__gte=today - timedelta(days=7)) & ~Q(assessments__isnull=False)),
    )
    upcoming_bookings    = booking_counts['upcoming'] or 0
    pending_assessments  = booking_counts['pending_assess'] or 0
    upcoming_sessions    = ScheduleBlock.objects.filter(coach=coach, date__gte=today, status='available').count()

    has_outstanding = upcoming_bookings > 0 or upcoming_sessions > 0 or pending_assessments > 0

    if request.method == 'POST':
        try:
            # Update user info
            coach.user.first_name = request.POST.get('first_name')
            coach.user.last_name = request.POST.get('last_name')
            coach.user.email = request.POST.get('email')
            coach.user.save()

            # Update coach profile — keep in sync with coach portal edit_profile
            coach.slug             = request.POST.get('slug', coach.user.username)
            coach.tagline          = request.POST.get('tagline', '')[:200]
            coach.bio              = request.POST.get('bio', '')
            coach.full_bio         = request.POST.get('full_bio', '')
            coach.specializations  = request.POST.get('specializations', '')
            coach.certifications   = request.POST.get('certifications', '')
            coach.experience_years = int(request.POST.get('experience_years', 0) or 0)
            coach.coaching_philosophy = request.POST.get('coaching_philosophy', '')
            coach.achievements     = request.POST.get('achievements', '')
            coach.hourly_rate      = request.POST.get('hourly_rate', 0)
            coach.instagram_url    = request.POST.get('instagram_url', '')
            coach.facebook_url     = request.POST.get('facebook_url', '')
            coach.twitter_url      = request.POST.get('twitter_url', '')
            coach.linkedin_url     = request.POST.get('linkedin_url', '')
            coach.youtube_url      = request.POST.get('youtube_url', '')
            coach.personal_website = request.POST.get('personal_website', '')
            coach.is_active        = request.POST.get('is_active') == 'on'
            coach.profile_enabled  = request.POST.get('profile_enabled') == 'on'

            # Photo upload
            if 'photo' in request.FILES:
                coach.photo = request.FILES['photo']
            elif request.POST.get('clear_photo'):
                coach.photo = None

            # Gallery images
            for i in (1, 2, 3):
                key = f'gallery_image_{i}'
                if key in request.FILES:
                    setattr(coach, key, request.FILES[key])
                elif request.POST.get(f'clear_gallery_{i}'):
                    setattr(coach, key, None)

            coach.save()

            messages.success(request, f'Coach "{coach.user.first_name}" updated successfully!')
            return redirect('owner_coaches')
        except Exception as e:
            messages.error(request, f'Error updating coach: {str(e)}')

    context = {
        'coach': coach,
        'editing': True,
        'upcoming_bookings': upcoming_bookings,
        'upcoming_sessions': upcoming_sessions,
        'pending_assessments': pending_assessments,
        'has_outstanding': has_outstanding,
    }
    return render(request, 'owner/coach_form.html', context)


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_coach_delete(request, pk):
    """Delete or deactivate a coach."""
    from django.shortcuts import get_object_or_404

    coach = get_object_or_404(Coach, pk=pk)
    today = timezone.localdate()
    permanent = request.POST.get('permanent') == 'true'

    if permanent:
        # Check for outstanding activities before permanent deletion
        upcoming_bookings = Booking.objects.filter(
            coach=coach,
            scheduled_date__gte=today,
            status__in=['pending', 'confirmed']
        ).count()

        upcoming_sessions = ScheduleBlock.objects.filter(
            coach=coach,
            date__gte=today,
            status='available'
        ).count()

        if upcoming_bookings > 0 or upcoming_sessions > 0:
            messages.error(request, f'Cannot delete coach with {upcoming_bookings} upcoming bookings and {upcoming_sessions} scheduled sessions. Please resolve these first.')
            return redirect('owner_coach_edit', pk=pk)

        # Permanent deletion
        coach_name = f"{coach.user.first_name} {coach.user.last_name}"
        user = coach.user

        # Remove from Coach group
        from django.contrib.auth.models import Group
        coach_group = Group.objects.filter(name='Coach').first()
        if coach_group:
            user.groups.remove(coach_group)

        # Delete coach profile
        coach.delete()

        # Delete user account
        user.delete()

        messages.success(request, f'Coach "{coach_name}" has been permanently deleted.')
    else:
        # Just deactivate
        coach.is_active = False
        coach.save()
        messages.success(request, f'Coach "{coach.user.first_name}" has been deactivated.')

    return redirect('owner_coaches')


@login_required
@user_passes_test(is_owner)
def owner_coach_schedule(request, pk):
    """Manage a coach's schedule blocks."""
    from django.shortcuts import get_object_or_404

    coach = get_object_or_404(Coach, pk=pk)
    today = timezone.localdate()

    # Get upcoming schedule blocks
    schedule_blocks = ScheduleBlock.objects.filter(
        coach=coach,
        date__gte=today
    ).order_by('date', 'start_time')[:30]

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_block':
            try:
                ScheduleBlock.objects.create(
                    coach=coach,
                    date=request.POST.get('date'),
                    start_time=request.POST.get('start_time'),
                    end_time=request.POST.get('end_time'),
                    session_type=request.POST.get('session_type', 'private'),
                    max_participants=request.POST.get('max_participants', 1),
                    notes=request.POST.get('notes', ''),
                )
                messages.success(request, 'Schedule block added successfully!')
            except Exception as e:
                messages.error(request, f'Error adding block: {str(e)}')

        elif action == 'delete_block':
            block_id = request.POST.get('block_id')
            try:
                block = ScheduleBlock.objects.get(pk=block_id, coach=coach)
                if block.current_participants == 0:
                    block.delete()
                    messages.success(request, 'Schedule block deleted.')
                else:
                    messages.error(request, 'Cannot delete block with existing bookings.')
            except ScheduleBlock.DoesNotExist:
                messages.error(request, 'Block not found.')

        return redirect('owner_coach_schedule', pk=pk)

    context = {
        'coach': coach,
        'schedule_blocks': schedule_blocks,
        'session_types': ScheduleBlock.SESSION_TYPE_CHOICES,
    }
    return render(request, 'owner/coach_schedule.html', context)


# ============================================================================
# BOOKING/SESSION MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_bookings(request):
    """List all bookings with filters."""
    from clients.models import Package

    today = timezone.localdate()
    status_filter       = request.GET.get('status', '')
    coach_filter        = request.GET.get('coach', '')
    date_filter         = request.GET.get('date', '')
    session_type_filter = request.GET.get('session_type', '')
    package_type_filter = request.GET.get('package_type', '')

    bookings = Booking.objects.select_related(
        'client__user', 'player', 'coach__user', 'session_type',
        'client_package__package',
    ).order_by('-scheduled_date', '-scheduled_time')

    if status_filter:
        bookings = bookings.filter(status=status_filter)
    if coach_filter:
        bookings = bookings.filter(coach_id=coach_filter)
    if date_filter:
        bookings = bookings.filter(scheduled_date=date_filter)
    if session_type_filter:
        bookings = bookings.filter(session_type_id=session_type_filter)
    if package_type_filter:
        bookings = bookings.filter(client_package__package__package_type=package_type_filter)

    # Limit to 200 recent bookings
    bookings = bookings[:200]

    context = {
        'bookings': bookings,
        'coaches': Coach.objects.filter(is_active=True),
        'session_types': SessionType.objects.filter(is_active=True).order_by('name'),
        'status_choices': Booking.STATUS_CHOICES,
        'package_type_choices': Package.PACKAGE_TYPE_CHOICES,
        'status_filter': status_filter,
        'coach_filter': coach_filter,
        'date_filter': date_filter,
        'session_type_filter': session_type_filter,
        'package_type_filter': package_type_filter,
    }
    return render(request, 'owner/bookings.html', context)


@login_required
@user_passes_test(is_owner)
def owner_booking_detail(request, pk):
    """View and manage a specific booking."""
    from django.shortcuts import get_object_or_404

    booking = get_object_or_404(
        Booking.objects.select_related('client__user', 'player', 'coach__user', 'session_type', 'client_package'),
        pk=pk
    )

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'confirm':
            try:
                booking.confirm()
                messages.success(request, 'Booking confirmed!')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')

        elif action == 'cancel':
            reason = request.POST.get('reason', 'other')
            notes = request.POST.get('notes', '')
            try:
                booking.status = 'cancelled'
                booking.cancellation_reason = reason
                booking.cancellation_notes = notes
                booking.cancelled_at = timezone.now()
                booking.save()
                messages.success(request, 'Booking cancelled.')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')

        elif action == 'complete':
            try:
                booking.status = 'completed'
                booking.completed_at = timezone.now()
                booking.save()
                messages.success(request, 'Booking marked as completed.')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')

        elif action == 'no_show':
            booking.status = 'no_show'
            booking.save()
            messages.success(request, 'Booking marked as no-show.')

        return redirect('owner_booking_detail', pk=pk)

    context = {
        'booking': booking,
        'cancellation_reasons': Booking.CANCELLATION_REASON_CHOICES,
    }
    return render(request, 'owner/booking_detail.html', context)


# ============================================================================
# CLIENT/PLAYER MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_clients(request):
    """List all clients with their players - only users in Client group."""
    from clients.models import ClientPackage
    from django.contrib.auth.models import Group

    # Only show clients who are in the Client group (not coaches with client profiles)
    client_group = Group.objects.filter(name='Client').first()
    if client_group:
        client_user_ids = client_group.user_set.values_list('id', flat=True)
        clients = Client.objects.filter(user_id__in=client_user_ids).select_related('user').annotate(
            player_count=Count('players', distinct=True),
            active_packages=Count('packages', filter=Q(packages__status='active'), distinct=True),
            total_bookings=Count('bookings', distinct=True)
        ).order_by('-created_at')[:100]
    else:
        clients = Client.objects.none()

    context = {
        'clients': clients,
    }
    return render(request, 'owner/clients.html', context)


@login_required
@user_passes_test(is_owner)
def owner_client_detail(request, pk):
    """View a client's details including players and bookings."""
    from django.shortcuts import get_object_or_404
    from clients.models import ClientPackage, RentalService

    client = get_object_or_404(Client.objects.select_related('user').prefetch_related('allowed_services'), pk=pk)
    players = Player.objects.filter(client=client, is_active=True)
    packages = ClientPackage.objects.filter(client=client).select_related('package')
    recent_bookings = Booking.objects.filter(client=client).select_related('player', 'coach__user')[:20]
    all_services = RentalService.objects.filter(is_active=True)

    context = {
        'client': client,
        'players': players,
        'packages': packages,
        'recent_bookings': recent_bookings,
        'all_services': all_services,
        'allowed_service_ids': list(client.allowed_services.values_list('id', flat=True)),
    }
    return render(request, 'owner/client_detail.html', context)


@login_required
@user_passes_test(is_owner)
def owner_client_approve(request, pk):
    """Approve a coach or renter client with term dates and allowed services."""
    from django.shortcuts import get_object_or_404
    from django.utils import timezone
    from clients.models import RentalService, Notification

    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        term_start_str = request.POST.get('term_start', '').strip()
        term_end_str   = request.POST.get('term_end', '').strip()
        service_ids    = request.POST.getlist('allowed_services')
        notes          = request.POST.get('approval_notes', '').strip()

        from datetime import datetime
        def parse_dt(s):
            for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
                try:
                    return datetime.strptime(s, fmt)
                except ValueError:
                    continue
            return None

        client.approval_status = 'approved'
        client.approved_by     = request.user
        client.approved_at     = timezone.now()
        client.approval_notes  = notes
        client.term_start      = parse_dt(term_start_str) if term_start_str else None
        client.term_end        = parse_dt(term_end_str) if term_end_str else None
        client.save()
        client.allowed_services.set(RentalService.objects.filter(id__in=service_ids))

        # Notify client
        Notification.objects.create(
            client=client,
            notification_type='promotional',
            title='Your access has been approved!',
            message=f'Your {client.get_client_type_display()} access has been approved by APC.'
                    + (f'\nTerm: {client.term_start.strftime("%b %d, %Y") if client.term_start else "Immediate"}'
                       + (f' → {client.term_end.strftime("%b %d, %Y")}' if client.term_end else '') if client.term_start or client.term_end else ''),
            method='email',
        ).send()

        messages.success(request, f'{client} approved successfully.')
    return redirect('owner_client_detail', pk=pk)


@login_required
@user_passes_test(is_owner)
def owner_client_reject(request, pk):
    """Reject a coach or renter client access request."""
    from django.shortcuts import get_object_or_404
    from django.utils import timezone
    from clients.models import Notification

    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        notes = request.POST.get('rejection_notes', '').strip()
        client.approval_status = 'rejected'
        client.rejected_at     = timezone.now()
        client.approval_notes  = notes
        client.save()

        Notification.objects.create(
            client=client,
            notification_type='promotional',
            title='Access request not approved',
            message=f'Your {client.get_client_type_display()} access request could not be approved at this time.'
                    + (f'\n\nNote from APC: {notes}' if notes else '')
                    + '\n\nPlease contact us if you have questions.',
            method='email',
        ).send()

        messages.warning(request, f'{client} access rejected.')
    return redirect('owner_client_detail', pk=pk)


@login_required
@user_passes_test(is_owner)
def owner_players(request):
    """List all players with session-type / package filters and CSV/XLSX/PDF export."""
    from clients.models import ClientPackage
    from bookings.models import Booking, SessionType as ST

    # ── Build base queryset ──────────────────────────────────────────────────
    qs = Player.objects.select_related('client__user').annotate(
        total_bookings=Count('bookings', distinct=True),
        total_assessments=Count('assessments', distinct=True),
    )

    # ── Filters ──────────────────────────────────────────────────────────────
    session_type_id = request.GET.get('session_type')
    package_id      = request.GET.get('package')
    search          = request.GET.get('q', '').strip()

    if session_type_id:
        qs = qs.filter(bookings__session_type_id=session_type_id).distinct()
    if package_id:
        qs = qs.filter(client__packages__package_id=package_id).distinct()
    if search:
        qs = qs.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(client__user__email__icontains=search)
        )

    players = qs.order_by('first_name', 'last_name')

    # ── Export ───────────────────────────────────────────────────────────────
    export = request.GET.get('export')
    if export in ('csv', 'xlsx', 'pdf'):
        # Load full data for export (no limit)
        rows = []
        for p in players.prefetch_related('client__packages__package'):
            active_pkgs = ', '.join(
                cp.package.name
                for cp in p.client.packages.filter(status='active').select_related('package')
            )
            rows.append({
                'First Name':      p.first_name,
                'Last Name':       p.last_name,
                'Birth Year':      p.birth_year,
                'Age Group':       p.age_group,
                'Gender':          p.get_gender_display(),
                'Position':        p.get_primary_position_display() or '',
                'Skill Level':     p.get_skill_level_display(),
                'Soccer Club':     p.soccer_club,
                'Team Name':       p.team_name,
                'Jersey Size':     p.get_jersey_size_display() if p.jersey_size else '',
                'National Team':   p.favorite_national_team,
                'Club Team':       p.favorite_club_team,
                'Parent/Guardian': p.client.user.get_full_name() or p.client.user.username,
                'Parent Email':    p.client.user.email,
                'Parent Phone':    p.client.phone,
                'Total Bookings':  p.total_bookings,
                'Total Assessments': p.total_assessments,
                'Active Packages': active_pkgs,
                'Notes':           p.notes,
            })

        headers = list(rows[0].keys()) if rows else []

        if export == 'csv':
            import csv
            from django.http import HttpResponse
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="players.csv"'
            writer = csv.DictWriter(response, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
            return response

        elif export == 'xlsx':
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse

            wb = Workbook()
            ws = wb.active
            ws.title = 'Players'

            # Header row — styled
            header_fill = PatternFill('solid', fgColor='6366F1')
            header_font = Font(bold=True, color='FFFFFF')
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(h) + 4)

            # Data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row[h])

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="players.xlsx"'
            return response

        elif export == 'pdf':
            # Render a print-optimized HTML page — browser handles PDF conversion
            context = {'players': players, 'rows': rows, 'headers': headers,
                       'total': players.count()}
            return render(request, 'owner/players_print.html', context)

    # ── Normal page view (limit to 500 for performance) ──────────────────────
    filter_session_types = ST.objects.filter(is_active=True).order_by('name')
    from clients.models import Package
    filter_packages = Package.objects.filter(is_active=True).order_by('name')

    context = {
        'players':             players[:500],
        'total':               players.count(),
        'filter_session_types': filter_session_types,
        'filter_packages':     filter_packages,
        'selected_session':    session_type_id or '',
        'selected_package':    package_id or '',
        'search':              search,
    }
    return render(request, 'owner/players.html', context)


@login_required
@user_passes_test(is_owner)
def owner_player_detail(request, pk):
    """View a single player's profile, bookings, packages, and assessments."""
    player = get_object_or_404(Player.objects.select_related('client__user', 'team'), pk=pk)
    bookings = Booking.objects.filter(player=player).select_related('coach__user', 'session_type').order_by('-scheduled_date')[:30]
    from clients.models import ClientPackage
    packages = ClientPackage.objects.filter(client=player.client).select_related('package').order_by('-purchase_date')
    assessments = PlayerAssessment.objects.filter(player=player).select_related('coach__user').order_by('-assessment_date')[:10]
    context = {
        'player': player,
        'bookings': bookings,
        'packages': packages,
        'assessments': assessments,
    }
    return render(request, 'owner/player_detail.html', context)


# ============================================================================
# SESSION TYPE MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_session_types(request):
    """Manage session types."""
    base_st = SessionType.objects.annotate(total_bookings=Count('bookings'))
    session_types          = base_st.filter(is_active=True).order_by('name')
    archived_session_types = base_st.filter(is_active=False).order_by('name')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            try:
                st = SessionType.objects.create(
                    name=request.POST.get('name'),
                    description=request.POST.get('description', ''),
                    session_format=request.POST.get('session_format', 'private'),
                    duration_minutes=request.POST.get('duration_minutes', 60),
                    price=request.POST.get('price'),
                    max_participants=request.POST.get('max_participants', 1),
                    color=request.POST.get('color', '#2ecc71'),
                    is_active=request.POST.get('is_active') == 'on',
                    requires_package=request.POST.get('requires_package') == 'on',
                    allow_package=request.POST.get('allow_package') == 'on',
                    show_as_event=request.POST.get('show_as_event') == 'on',
                    show_as_program=request.POST.get('show_as_program') == 'on',
                    start_times=' '.join(t for t in request.POST.getlist('start_times') if t),
                    location=request.POST.get('location', ''),
                    age_group=request.POST.get('age_group', ''),
                    days_of_week=','.join(request.POST.getlist('days_of_week')),
                    # Clinic/Camp fields
                    start_date=request.POST.get('start_date') or None,
                    end_date=request.POST.get('end_date') or None,
                    min_age=request.POST.get('min_age') or None,
                    max_age=request.POST.get('max_age') or None,
                )
                pkg_ids = request.POST.getlist('linked_packages')
                if pkg_ids:
                    from clients.models import Package as Pkg
                    st.linked_packages.set(Pkg.objects.filter(pk__in=pkg_ids))
                messages.success(request, 'Session type created!')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')

        elif action == 'toggle':
            st_id = request.POST.get('session_type_id')
            try:
                st = SessionType.objects.get(pk=st_id)
                st.is_active = not st.is_active
                st.save()
                messages.success(request, f'Session type {"activated" if st.is_active else "deactivated"}.')
            except SessionType.DoesNotExist:
                messages.error(request, 'Session type not found.')

        return redirect('owner_session_types')

    from clients.models import Package
    context = {
        'session_types': session_types,
        'archived_session_types': archived_session_types,
        'format_choices': SessionType.SESSION_FORMAT_CHOICES,
        'days_of_week_choices': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
        'packages': Package.objects.filter(is_active=True).order_by('name'),
        'archived_linked_packages': [],
        'linked_package_ids': [],
    }
    return render(request, 'owner/session_types.html', context)


@login_required
@user_passes_test(is_owner)
def owner_session_type_edit(request, pk):
    """Edit a session type."""
    from django.shortcuts import get_object_or_404

    session_type = get_object_or_404(SessionType, pk=pk)

    if request.method == 'POST':
        try:
            session_type.name = request.POST.get('name')
            session_type.description = request.POST.get('description', '')
            session_type.session_format = request.POST.get('session_format', 'private')
            session_type.duration_minutes = request.POST.get('duration_minutes', 60)
            session_type.price = request.POST.get('price')
            drop_in = request.POST.get('drop_in_price', '').strip()
            session_type.drop_in_price = drop_in if drop_in else None
            session_type.max_participants = request.POST.get('max_participants', 1)
            session_type.color = request.POST.get('color', '#2ecc71')
            session_type.is_active = request.POST.get('is_active') == 'on'
            session_type.requires_package = request.POST.get('requires_package') == 'on'
            session_type.allow_package    = request.POST.get('allow_package') == 'on'
            session_type.show_as_event = request.POST.get('show_as_event') == 'on'
            session_type.show_as_program = request.POST.get('show_as_program') == 'on'
            # Poster image
            if request.POST.get('clear_poster_image') and session_type.poster_image:
                session_type.poster_image.delete(save=False)
                session_type.poster_image = None
            elif 'poster_image' in request.FILES:
                new_poster = request.FILES['poster_image']
                from clients.utils import validate_photo
                err = validate_photo(new_poster)
                if err:
                    raise ValueError(err)
                if session_type.poster_image:
                    session_type.poster_image.delete(save=False)
                session_type.poster_image = new_poster
            # Carousel CTA + order
            session_type.event_cta_text = request.POST.get('event_cta_text', '').strip()
            session_type.event_cta_url = request.POST.get('event_cta_url', '').strip()
            try:
                session_type.event_display_order = int(request.POST.get('event_display_order', 0))
            except (ValueError, TypeError):
                session_type.event_display_order = 0
            session_type.start_times = ' '.join(t for t in request.POST.getlist('start_times') if t)
            session_type.weekend_start_times = ' '.join(t for t in request.POST.getlist('weekend_start_times') if t)
            session_type.location = request.POST.get('location', '')
            session_type.age_group = request.POST.get('age_group', '')
            session_type.days_of_week = ','.join(request.POST.getlist('days_of_week'))
            # Clinic/Camp fields
            session_type.start_date = request.POST.get('start_date') or None
            session_type.end_date = request.POST.get('end_date') or None
            session_type.min_age = request.POST.get('min_age') or None
            session_type.max_age = request.POST.get('max_age') or None
            # Per-day/time capacity rules: fields named cap_<Day>_<HH:MM>
            day_capacities = {}
            for key, val in request.POST.items():
                if key.startswith('cap_') and val.strip():
                    try:
                        day_capacities[key[4:]] = int(val)
                    except ValueError:
                        pass
            session_type.day_capacities = day_capacities
            session_type.save()
            pkg_ids = request.POST.getlist('linked_packages')
            from clients.models import Package as Pkg
            session_type.linked_packages.set(Pkg.objects.filter(pk__in=pkg_ids))
            messages.success(request, f'Session type "{session_type.name}" updated!')
            return redirect('owner_session_types')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    from clients.models import Package as Pkg
    linked_ids = list(session_type.linked_packages.values_list('pk', flat=True))
    # Purchasable active packages shown in selector
    purchasable_pkgs = Pkg.objects.filter(is_active=True, is_purchasable=True).order_by('name')
    # Also show active-but-not-purchasable (e.g. spring packages) in selector — clients may still hold them
    active_pkgs = Pkg.objects.filter(is_active=True).order_by('name')
    # Archived packages already linked → shown as read-only (preserve link)
    archived_linked = Pkg.objects.filter(pk__in=linked_ids, is_active=False)
    context = {
        'session_type': session_type,
        'format_choices': SessionType.SESSION_FORMAT_CHOICES,
        'packages': active_pkgs,
        'archived_linked_packages': archived_linked,
        'linked_package_ids': linked_ids,
        'days_of_week_choices': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
    }
    return render(request, 'owner/session_type_form.html', context)


@login_required
@user_passes_test(is_owner)
def owner_session_type_apply_capacities(request, pk):
    """AJAX: bulk-update ScheduleBlock.max_participants for blocks linked to this session type."""
    import json
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from coaches.models import ScheduleBlock

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    session_type = get_object_or_404(SessionType, pk=pk)

    try:
        data = json.loads(request.body)
        capacities = data.get('capacities', {})  # {"Mon_17:00": 20, ...}
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    if not capacities:
        return JsonResponse({'error': 'No capacity rules provided'}, status=400)

    # Day abbr → weekday integer (Monday=0)
    day_map = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}

    blocks = ScheduleBlock.objects.filter(catalog_session_types=session_type)
    updated = 0

    for block in blocks:
        day_abbr = list(day_map.keys())[block.date.weekday()]
        time_str = block.start_time.strftime('%H:%M')
        key = f"{day_abbr}_{time_str}"
        if key in capacities:
            new_cap = int(capacities[key])
            if block.max_participants != new_cap:
                block.max_participants = new_cap
                block.save(update_fields=['max_participants'])
                updated += 1

    # Also save capacities to the session type
    session_type.day_capacities = {k: int(v) for k, v in capacities.items()}
    session_type.save(update_fields=['day_capacities'])

    return JsonResponse({'updated': updated})


@login_required
@user_passes_test(is_owner)
def owner_session_type_roster(request, pk):
    """Capacity roster: bookings vs max per day/time for a session type."""
    from django.shortcuts import get_object_or_404
    from coaches.models import ScheduleBlock
    from django.db.models import Count, Q

    session_type = get_object_or_404(SessionType, pk=pk)

    # All schedule blocks linked to this session type
    blocks = (
        ScheduleBlock.objects
        .filter(catalog_session_types=session_type)
        .select_related('coach__user')
        .order_by('date', 'start_time')
    )

    # Pre-fetch booking counts keyed by (coach_id, date, start_time)
    booking_counts = {}
    for b in (
        Booking.objects
        .filter(session_type=session_type, status__in=['pending', 'confirmed'])
        .values('coach_id', 'scheduled_date', 'scheduled_time')
        .annotate(cnt=Count('id'))
    ):
        booking_counts[(b['coach_id'], b['scheduled_date'], b['scheduled_time'])] = b['cnt']

    # Build roster rows with fill metrics
    roster = []
    for block in blocks:
        booked = booking_counts.get(
            (block.coach_id, block.date, block.start_time), 0
        )
        capacity = session_type.get_capacity(
            ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][block.date.weekday()],
            block.start_time.strftime('%H:%M'),
        )
        pct = round(booked / capacity * 100) if capacity else 0
        roster.append({
            'block':    block,
            'booked':   booked,
            'capacity': capacity,
            'pct':      pct,
            'status':   'full' if pct >= 100 else ('warning' if pct >= 70 else 'ok'),
        })

    context = {
        'session_type': session_type,
        'roster':       roster,
        'total_booked': sum(r['booked'] for r in roster),
        'total_capacity': sum(r['capacity'] for r in roster),
    }
    return render(request, 'owner/session_type_roster.html', context)


# ============================================================================
# TEAM MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_teams(request):
    """List all teams with stats."""
    from clients.models import Team, ClientPackage
    from django.db.models import Count, Q

    teams = Team.objects.filter(is_active=True).annotate(
        active_player_count=Count('players', filter=Q(players__is_active=True)),
        coach_count=Count('coaches')
    ).order_by('age_group', 'name')

    # Calculate stats
    total_teams = teams.count()
    total_players = Player.objects.filter(is_active=True, team__is_active=True).count()
    total_coaches = Coach.objects.filter(teams__is_active=True).distinct().count()
    
    # Count active team packages
    active_packages = ClientPackage.objects.filter(
        status='active',
        package__package_type='team'
    ).count()

    context = {
        'teams': teams,
        'total_teams': total_teams,
        'total_players': total_players,
        'total_coaches': total_coaches,
        'active_packages': active_packages,
    }
    return render(request, 'owner/teams.html', context)


@login_required
@user_passes_test(is_owner)
def owner_team_detail(request, pk):
    """Show detailed team info with roster, coaches, and bookings."""
    from clients.models import Team
    from django.shortcuts import get_object_or_404
    from datetime import timedelta

    team = get_object_or_404(Team.objects.select_related('manager__user'), pk=pk)
    today = timezone.localdate()

    # Get team players with details
    players = Player.objects.filter(team=team, is_active=True).select_related('client__user')

    # Get assigned coaches
    coaches = team.coaches.all().select_related('user')

    # Get recent and upcoming bookings for team players
    recent_bookings = Booking.objects.filter(
        player__team=team
    ).select_related('player', 'coach__user', 'session_type').order_by('-scheduled_date')[:10]

    upcoming_bookings = Booking.objects.filter(
        player__team=team,
        scheduled_date__gte=today,
        status__in=['pending', 'confirmed']
    ).select_related('player', 'coach__user', 'session_type').order_by('scheduled_date')[:10]

    # Get package usage for team players
    from clients.models import ClientPackage
    team_packages = ClientPackage.objects.filter(
        player__team=team
    ).select_related('package', 'player').order_by('-purchase_date')[:10]

    context = {
        'team': team,
        'players': players,
        'coaches': coaches,
        'recent_bookings': recent_bookings,
        'upcoming_bookings': upcoming_bookings,
        'team_packages': team_packages,
        'player_count': players.count(),
        'coach_count': coaches.count(),
    }
    return render(request, 'owner/team_detail.html', context)


@login_required
@user_passes_test(is_owner)
def owner_team_players(request, team_id):
    """View all players on a specific team."""
    from clients.models import Team
    from django.shortcuts import get_object_or_404

    team = get_object_or_404(Team.objects.select_related('manager__user'), pk=team_id)
    
    players = Player.objects.filter(team=team, is_active=True).select_related('client__user').annotate(
        total_bookings=Count('bookings'),
        total_assessments=Count('assessments')
    ).order_by('first_name', 'last_name')

    context = {
        'team': team,
        'players': players,
    }
    return render(request, 'owner/team_players.html', context)


@login_required
@user_passes_test(is_owner)
def owner_team_bookings(request, team_id):
    """View all bookings for a specific team."""
    from clients.models import Team
    from django.shortcuts import get_object_or_404

    team = get_object_or_404(Team.objects.select_related('manager__user'), pk=team_id)
    today = timezone.localdate()

    # Get date filters from request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    bookings = Booking.objects.filter(player__team=team).select_related(
        'player', 'coach__user', 'session_type'
    ).order_by('-scheduled_date', '-scheduled_time')

    if date_from:
        bookings = bookings.filter(scheduled_date__gte=date_from)
    if date_to:
        bookings = bookings.filter(scheduled_date__lte=date_to)

    # Calculate summary stats — single aggregate instead of 4 count queries
    booking_stats = bookings.aggregate(
        total_bookings=Count('id'),
        completed_bookings=Count('id', filter=Q(status='completed')),
        upcoming_bookings=Count('id', filter=Q(scheduled_date__gte=today, status__in=['pending', 'confirmed'])),
        cancelled_bookings=Count('id', filter=Q(status='cancelled')),
    )
    total_bookings     = booking_stats['total_bookings'] or 0
    completed_bookings = booking_stats['completed_bookings'] or 0
    upcoming_bookings  = booking_stats['upcoming_bookings'] or 0
    cancelled_bookings = booking_stats['cancelled_bookings'] or 0

    context = {
        'team': team,
        'bookings': bookings[:100],
        'total_bookings': total_bookings,
        'completed_bookings': completed_bookings,
        'upcoming_bookings': upcoming_bookings,
        'cancelled_bookings': cancelled_bookings,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'owner/team_bookings.html', context)


# ============================================================================
# FIELD RENTAL MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_field_slots(request):
    """List and create rental slots. Also handles service catalog CRUD."""
    from clients.models import FieldRentalSlot, RentalService, Notification, Client
    from datetime import datetime as dt

    today = timezone.localdate()

    # --- Service catalog actions (merged from owner_services) ---
    action = request.POST.get('action', '')
    if request.method == 'POST' and action == 'service_create':
        try:
            RentalService.objects.create(
                name=request.POST['name'],
                service_type=request.POST['service_type'],
                description=request.POST.get('description', ''),
                capacity=request.POST.get('capacity') or None,
                price=request.POST['price'],
                pricing_type=request.POST.get('pricing_type', 'flat'),
                requires_approval=request.POST.get('requires_approval') == 'on',
                is_active=True,
            )
            messages.success(request, 'Service added.')
        except Exception as e:
            messages.error(request, f'Error creating service: {e}')
        return redirect('owner_field_slots')

    if request.method == 'POST' and action == 'service_save':
        svc = get_object_or_404(RentalService, pk=request.POST.get('service_id'))
        try:
            svc.name = request.POST['name']
            svc.service_type = request.POST['service_type']
            svc.description = request.POST.get('description', '')
            svc.capacity = request.POST.get('capacity') or None
            svc.price = request.POST['price']
            svc.pricing_type = request.POST.get('pricing_type', 'flat')
            svc.requires_approval = request.POST.get('requires_approval') == 'on'
            svc.is_active = request.POST.get('is_active') == 'on'
            svc.save()
            messages.success(request, f'"{svc.name}" updated.')
        except Exception as e:
            messages.error(request, f'Error updating service: {e}')
        return redirect('owner_field_slots')

    if request.method == 'POST' and action == 'service_delete':
        svc = get_object_or_404(RentalService, pk=request.POST.get('service_id'))
        active_slots = svc.slots.filter(status__in=['pending_approval', 'booked']).count()
        if active_slots:
            messages.error(request, f'Cannot delete: {active_slots} active slot(s) use this service.')
        else:
            svc.delete()
            messages.success(request, 'Service deleted.')
        return redirect('owner_field_slots')

    if request.method == 'POST' and action == 'add':
        try:
            start_str = request.POST.get('start_time', '')
            end_str    = request.POST.get('end_time', '')
            start_t    = dt.strptime(start_str, '%H:%M').time()
            end_t      = dt.strptime(end_str,   '%H:%M').time()
            duration   = int((dt.combine(today, end_t) - dt.combine(today, start_t)).seconds / 60)
            service_id = request.POST.get('service_id') or None
            service    = RentalService.objects.get(pk=service_id) if service_id else None

            slot = FieldRentalSlot.objects.create(
                date=request.POST.get('date'),
                start_time=start_t,
                end_time=end_t,
                duration_minutes=duration,
                price=request.POST.get('price', 0),
                title=request.POST.get('title', ''),
                notes=request.POST.get('notes', ''),
                service=service,
            )
            if slot.has_conflicting_schedule_blocks:
                messages.warning(request,
                    f'Slot created, but existing coach schedule blocks overlap this time. '
                    f'Those blocks will be blocked from new bookings once a field rental is active.')
            else:
                messages.success(request, f'Field rental slot created for {slot.date}.')
        except Exception as e:
            messages.error(request, f'Error creating slot: {e}')
        return redirect('owner_field_slots')

    status_filter = request.GET.get('status', 'all')
    slots = FieldRentalSlot.objects.select_related('booked_by_client__user', 'booked_by_team')
    if status_filter != 'all':
        slots = slots.filter(status=status_filter)
    slots = slots.order_by('date', 'start_time')

    pending_slots = FieldRentalSlot.objects.filter(status='pending_approval').order_by('requested_at')
    revenue = FieldRentalSlot.objects.filter(
        status='booked', date__month=today.month, date__year=today.year, payment_status='paid'
    ).aggregate(total=Sum('amount_paid'))['total'] or 0

    from django.db.models import Prefetch
    status_filter_q = slots  # already filtered above

    # Services with their slots prefetched (filtered by status if set)
    slot_qs = FieldRentalSlot.objects.select_related(
        'booked_by_client__user', 'booked_by_team'
    ).order_by('date', 'start_time')
    if status_filter != 'all':
        slot_qs = slot_qs.filter(status=status_filter)

    services_with_slots = RentalService.objects.prefetch_related(
        Prefetch('slots', queryset=slot_qs, to_attr='filtered_slots')
    ).order_by('service_type', 'name')

    # Slots not linked to any service
    unlinked_slots = FieldRentalSlot.objects.filter(
        service__isnull=True
    ).select_related('booked_by_client__user', 'booked_by_team').order_by('date', 'start_time')
    if status_filter != 'all':
        unlinked_slots = unlinked_slots.filter(status=status_filter)

    context = {
        'slots': slots,
        'pending_slots': pending_slots,
        'status_filter': status_filter,
        'today': today,
        'statuses': [('available', 'Available'), ('pending_approval', 'Pending'), ('booked', 'Booked'), ('cancelled', 'Cancelled')],
        **FieldRentalSlot.objects.aggregate(
            available_count=Count('id', filter=Q(status='available', date__gte=today)),
            pending_count=Count('id', filter=Q(status='pending_approval')),
            booked_month=Count('id', filter=Q(status='booked', date__month=today.month, date__year=today.year)),
        ),
        'revenue_month':   revenue,
        'services':             RentalService.objects.filter(is_active=True).order_by('service_type', 'name'),
        'all_services':         RentalService.objects.all().order_by('service_type', 'name'),
        'services_with_slots':  services_with_slots,
        'unlinked_slots':       unlinked_slots,
        'service_type_choices': RentalService.SERVICE_TYPE_CHOICES,
        'pricing_type_choices': RentalService.PRICING_TYPE_CHOICES,
    }
    return render(request, 'owner/field_slots.html', context)


@login_required
@user_passes_test(is_owner)
def owner_field_slot_edit(request, pk):
    """Edit an available field rental slot."""
    from clients.models import FieldRentalSlot
    from datetime import datetime as dt

    slot = get_object_or_404(FieldRentalSlot, pk=pk)
    if slot.status != 'available':
        messages.error(request, 'Only available (unbooked) slots can be edited.')
        return redirect('owner_field_slots')

    if request.method == 'POST':
        try:
            today = timezone.localdate()
            start_t = dt.strptime(request.POST['start_time'], '%H:%M').time()
            end_t   = dt.strptime(request.POST['end_time'],   '%H:%M').time()
            slot.date             = request.POST['date']
            slot.start_time       = start_t
            slot.end_time         = end_t
            slot.duration_minutes = int((dt.combine(today, end_t) - dt.combine(today, start_t)).seconds / 60)
            slot.price            = request.POST['price']
            slot.title            = request.POST.get('title', '')
            slot.notes            = request.POST.get('notes', '')
            slot.save()
            messages.success(request, 'Slot updated.')
        except Exception as e:
            messages.error(request, f'Error updating slot: {e}')
    return redirect('owner_field_slots')


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_field_slot_approve(request, pk):
    """Approve a pending field rental request."""
    from clients.models import FieldRentalSlot, Notification

    slot = get_object_or_404(FieldRentalSlot, pk=pk, status='pending_approval')
    slot.status      = 'booked'
    slot.approved_at = timezone.now()
    slot.booked_at   = timezone.now()
    slot.save()

    # Notify requester
    if slot.booked_by_client:
        Notification.objects.create(
            client=slot.booked_by_client,
            notification_type='field_rental_approved',
            title='Field Rental Approved!',
            message=f'Your field rental request for {slot.date:%b %d, %Y} '
                    f'({slot.start_time:%I:%M %p}–{slot.end_time:%I:%M %p}) has been approved.',
            method='email',
        )

    messages.success(request, f'Field rental approved for {slot.requester_name}.')
    return redirect('owner_field_slots')


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_field_slot_reject(request, pk):
    """Reject a pending field rental request."""
    from clients.models import FieldRentalSlot, Notification

    slot = get_object_or_404(FieldRentalSlot, pk=pk, status='pending_approval')
    reason = request.POST.get('rejection_reason', 'No reason provided.')

    requesting_client = slot.booked_by_client

    slot.status           = 'available'
    slot.rejection_reason = reason
    slot.rejected_at      = timezone.now()
    slot.booked_by_client = None
    slot.booked_by_team   = None
    slot.booker_type      = None
    slot.requested_at     = None
    slot.client_notes     = ''
    slot.save()

    if requesting_client:
        Notification.objects.create(
            client=requesting_client,
            notification_type='field_rental_rejected',
            title='Field Rental Not Approved',
            message=f'Your field rental request for {slot.date:%b %d, %Y} was not approved. Reason: {reason}',
            method='email',
        )

    messages.warning(request, 'Field rental request rejected and slot returned to available.')
    return redirect('owner_field_slots')


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_field_slot_cancel(request, pk):
    """Owner cancels a confirmed field rental booking."""
    from clients.models import FieldRentalSlot, Notification

    slot = get_object_or_404(FieldRentalSlot, pk=pk, status='booked')
    note = request.POST.get('cancellation_notes', '')
    requesting_client = slot.booked_by_client

    slot.status             = 'available'
    slot.cancellation_notes = note
    slot.cancelled_at       = timezone.now()
    slot.booked_by_client   = None
    slot.booked_by_team     = None
    slot.booker_type        = None
    slot.approved_at        = None
    slot.booked_at          = None
    slot.save()

    if requesting_client:
        Notification.objects.create(
            client=requesting_client,
            notification_type='field_rental_cancelled',
            title='Field Rental Cancelled',
            message=f'Your field rental on {slot.date:%b %d, %Y} '
                    f'({slot.start_time:%I:%M %p}–{slot.end_time:%I:%M %p}) has been cancelled by the owner.'
                    + (f' Note: {note}' if note else ''),
            method='email',
        )

    messages.warning(request, 'Field rental booking cancelled and slot returned to available.')
    return redirect('owner_field_slots')


@login_required
@user_passes_test(is_owner)
def owner_field_slot_conflict_check(request):
    """AJAX: check for ScheduleBlock conflicts and same-service slot conflicts."""
    from django.http import JsonResponse
    from clients.models import FieldRentalSlot

    date       = request.GET.get('date')
    start_time = request.GET.get('start_time')
    end_time   = request.GET.get('end_time')
    service_id = request.GET.get('service_id') or None
    exclude_pk = request.GET.get('exclude_pk') or None

    if not all([date, start_time, end_time]):
        return JsonResponse({'conflict': False, 'count': 0, 'blocks': [], 'service_conflicts': []})

    # Coach schedule block conflicts (relevant for field types)
    conflicts = ScheduleBlock.objects.filter(
        date=date, status__in=['available', 'booked']
    ).exclude(
        end_time__lte=start_time
    ).exclude(
        start_time__gte=end_time
    ).select_related('coach__user')

    blocks = [
        {
            'coach': f"{b.coach.user.first_name} {b.coach.user.last_name}".strip(),
            'start': str(b.start_time),
            'end':   str(b.end_time),
            'type':  b.get_session_type_display(),
        }
        for b in conflicts
    ]

    # Same-service slot conflicts
    service_conflict_list = []
    if service_id:
        svc_conflicts = FieldRentalSlot.check_service_blocked(
            service_id=service_id,
            date=date,
            start_time=start_time,
            end_time=end_time,
            exclude_pk=exclude_pk,
        ).select_related('booked_by_client__user', 'booked_by_team')
        service_conflict_list = [
            {
                'start':    str(s.start_time),
                'end':      str(s.end_time),
                'status':   s.status,
                'booker':   s.requester_name,
            }
            for s in svc_conflicts
        ]

    return JsonResponse({
        'conflict':          conflicts.exists(),
        'count':             conflicts.count(),
        'blocks':            blocks,
        'service_conflicts': service_conflict_list,
    })


# ============================================================================
# SERVICE CATALOG MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_services(request):
    """Redirect to rentals page — service catalog is now embedded there."""
    return redirect('owner_field_slots')


@login_required
@user_passes_test(is_owner)
def owner_service_edit(request, pk):
    """Edit an existing service catalog entry."""
    from clients.models import RentalService
    service = get_object_or_404(RentalService, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'delete':
            active_slots = service.slots.filter(status__in=['pending_approval', 'booked']).count()
            if active_slots:
                messages.error(request, f'Cannot delete: {active_slots} active slot(s) use this service.')
            else:
                service.delete()
                messages.success(request, 'Service deleted.')
            return redirect('owner_services')

        try:
            service.name             = request.POST['name']
            service.service_type     = request.POST['service_type']
            service.description      = request.POST.get('description', '')
            service.capacity         = request.POST.get('capacity') or None
            service.price            = request.POST['price']
            service.pricing_type     = request.POST.get('pricing_type', 'flat')
            service.requires_approval = request.POST.get('requires_approval') == 'on'
            service.is_active        = request.POST.get('is_active') == 'on'
            service.save()
            messages.success(request, f'"{service.name}" updated.')
        except Exception as e:
            messages.error(request, f'Error updating service: {e}')
        return redirect('owner_services')

    context = {
        'service': service,
        'service_type_choices': RentalService.SERVICE_TYPE_CHOICES,
        'pricing_type_choices': RentalService.PRICING_TYPE_CHOICES,
    }
    return render(request, 'owner/service_edit.html', context)


# ============================================================================
# FINANCE DASHBOARD
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_finances(request):
    """Revenue reporting, outstanding balances, and transaction history."""
    from clients.models import ClientPackage, Package
    from clients.models import FieldRentalSlot
    from payments.models import Payment
    from calendar import month_name
    from decimal import Decimal

    today = timezone.localdate()

    # --- Date range: default to current month, support ?month=M&year=Y ---
    try:
        view_month = int(request.GET.get('month', today.month))
        view_year  = int(request.GET.get('year',  today.year))
        if not (1 <= view_month <= 12):
            view_month = today.month
        if not (2000 <= view_year <= 2100):
            view_year = today.year
    except (ValueError, TypeError):
        view_month, view_year = today.month, today.year

    # Prev / next month navigation
    if view_month == 1:
        prev_month, prev_year = 12, view_year - 1
    else:
        prev_month, prev_year = view_month - 1, view_year
    if view_month == 12:
        next_month, next_year = 1, view_year + 1
    else:
        next_month, next_year = view_month + 1, view_year

    # ---- Revenue for selected month ----------------------------------------

    # Sessions paid directly
    session_revenue = Booking.objects.filter(
        scheduled_date__month=view_month,
        scheduled_date__year=view_year,
        payment_status='paid',
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')

    # Package revenue = actual amount charged via Stripe (not list price)
    package_revenue = Payment.objects.filter(
        status='succeeded',
        created_at__month=view_month,
        created_at__year=view_year,
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # Facility rentals paid
    rental_revenue = FieldRentalSlot.objects.filter(
        approved_at__month=view_month,
        approved_at__year=view_year,
        payment_status='paid',
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')

    total_revenue = session_revenue + package_revenue + rental_revenue

    # ---- Tax calculations --------------------------------------------------
    tax_rate = Decimal(str(getattr(settings, 'TAX_RATE', 0.0)))
    tax_amount = (total_revenue * tax_rate).quantize(Decimal('0.01'))
    revenue_after_tax = total_revenue - tax_amount
    tax_enabled = tax_rate > 0

    # ---- Stripe confirmed payments for this month (pre-wired for go-live) --
    stripe_confirmed = Payment.objects.filter(
        status='succeeded',
        created_at__month=view_month,
        created_at__year=view_year,
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    stripe_count = Payment.objects.filter(
        status='succeeded',
        created_at__month=view_month,
        created_at__year=view_year,
    ).count()

    # ---- 6-month trend -----------------------------------------------------
    monthly_trend = []
    for i in range(5, -1, -1):
        # Walk back i months from view_month/view_year
        m = view_month - i
        y = view_year
        while m <= 0:
            m += 12
            y -= 1

        s = Booking.objects.filter(
            scheduled_date__month=m, scheduled_date__year=y, payment_status='paid'
        ).aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')

        p = Payment.objects.filter(
            status='succeeded',
            created_at__month=m, created_at__year=y
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        r = FieldRentalSlot.objects.filter(
            approved_at__month=m, approved_at__year=y, payment_status='paid'
        ).aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')

        monthly_trend.append({
            'label': f"{month_name[m][:3]} {str(y)[2:]}",
            'sessions': float(s),
            'packages': float(p),
            'rentals':  float(r),
            'total':    float(s + p + r),
        })

    max_total = max((m['total'] for m in monthly_trend), default=1) or 1

    # ---- Outstanding balances (unpaid bookings) ----------------------------
    outstanding = Booking.objects.filter(
        payment_status='pending',
        status__in=['pending', 'confirmed'],
    ).select_related('client__user', 'player', 'session_type').order_by('scheduled_date')

    outstanding_total = outstanding.aggregate(
        total=Sum('session_type__price')
    )['total'] or Decimal('0')

    # ---- Recent transactions -----------------------------------------------
    recent_bookings = Booking.objects.filter(
        payment_status='paid',
    ).select_related('client__user', 'player', 'session_type').order_by('-scheduled_date')[:15]

    recent_packages = ClientPackage.objects.exclude(
        status='cancelled'
    ).select_related('client__user', 'package').order_by('-purchase_date')[:10]

    recent_rentals = FieldRentalSlot.objects.filter(
        payment_status='paid',
    ).select_related('booked_by_client__user', 'service').order_by('-approved_at')[:10]

    # ---- Package sales summary for the month --------------------------------
    package_breakdown = ClientPackage.objects.filter(
        purchase_date__month=view_month,
        purchase_date__year=view_year,
    ).exclude(status='cancelled').values(
        'package__name'
    ).annotate(
        count=Count('id'),
        revenue=Sum('package__price'),
    ).order_by('-revenue')

    context = {
        'today': today,
        'view_month': view_month,
        'view_year': view_year,
        'view_month_name': month_name[view_month],
        'prev_month': prev_month, 'prev_year': prev_year,
        'next_month': next_month, 'next_year': next_year,
        # Revenue totals
        'session_revenue':    session_revenue,
        'package_revenue':    package_revenue,
        'rental_revenue':     rental_revenue,
        'total_revenue':      total_revenue,
        # Tax
        'tax_rate':           tax_rate,
        'tax_rate_pct':       float(tax_rate * 100),
        'tax_amount':         tax_amount,
        'revenue_after_tax':  revenue_after_tax,
        'tax_enabled':        tax_enabled,
        # Stripe
        'stripe_confirmed':   stripe_confirmed,
        'stripe_count':       stripe_count,
        'stripe_live':        bool(settings.STRIPE_SECRET_KEY and settings.STRIPE_SECRET_KEY.startswith(('sk_live', 'rk_live'))),
        # Trend
        'monthly_trend':    monthly_trend,
        'max_total':        max_total,
        # Outstanding
        'outstanding':         outstanding[:20],
        'outstanding_total':   outstanding_total,
        'outstanding_count':   outstanding.count(),
        # Transactions
        'recent_bookings':  recent_bookings,
        'recent_packages':  recent_packages,
        'recent_rentals':   recent_rentals,
        # Package breakdown
        'package_breakdown': package_breakdown,
    }
    return render(request, 'owner/finances.html', context)


# ── Stripe Payments (owner portal) ───────────────────────────────────────────

@login_required
@user_passes_test(is_owner)
def owner_payments(request):
    """List all Stripe payment records."""
    from payments.models import Payment
    payments = Payment.objects.select_related('client__user').order_by('-created_at')[:100]
    context = {
        'payments': payments,
        'stripe_live': bool(settings.STRIPE_SECRET_KEY and settings.STRIPE_SECRET_KEY.startswith(('sk_live', 'rk_live'))),
    }
    return render(request, 'owner/payments.html', context)


@login_required
@user_passes_test(is_owner)
@require_POST
def owner_issue_refund(request, payment_id):
    """Issue a full or partial Stripe refund."""
    from payments.models import Payment
    from django.shortcuts import get_object_or_404

    if not settings.STRIPE_SECRET_KEY:
        messages.error(request, 'Stripe is not configured.')
        return redirect('owner_payments')

    import stripe
    stripe.api_key = settings.STRIPE_SECRET_KEY

    payment = get_object_or_404(Payment, pk=payment_id, status='succeeded')
    amount_str = request.POST.get('amount', '').strip()

    try:
        kwargs = {'payment_intent': payment.stripe_payment_intent_id}
        if amount_str:
            kwargs['amount'] = int(float(amount_str) * 100)
        stripe.Refund.create(**kwargs)
        messages.success(request, f'Refund initiated for {payment.client} — ${payment.amount}. Status updates via webhook.')
    except stripe.error.StripeError as e:
        messages.error(request, f'Refund failed: {e.user_message}')

    return redirect('owner_payments')


# ============================================================================
# CREDITS MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_credits(request):
    """Manage client credits — view, grant, and apply credits."""
    from django.utils import timezone as tz

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'grant':
            client_id = request.POST.get('client_id')
            amount = request.POST.get('amount', '').strip()
            credit_type = request.POST.get('credit_type', 'manual')
            notes = request.POST.get('notes', '')
            expires_str = request.POST.get('expires_at', '').strip()

            try:
                client = Client.objects.get(pk=client_id)
                from decimal import Decimal
                credit = ClientCredit.objects.create(
                    client=client,
                    amount=Decimal(amount),
                    credit_type=credit_type,
                    notes=notes,
                    expires_at=expires_str or None,
                    created_by=request.user,
                )
                messages.success(request, f'${credit.amount} credit granted to {client.user.get_full_name() or client.user.username}.')
            except Exception as e:
                messages.error(request, f'Error granting credit: {e}')

        elif action == 'cancel':
            credit_id = request.POST.get('credit_id')
            credit = get_object_or_404(ClientCredit, pk=credit_id)
            if credit.status == 'available':
                credit.status = 'cancelled'
                credit.save()
                messages.success(request, 'Credit cancelled.')
            else:
                messages.error(request, 'Only available credits can be cancelled.')

        return redirect('owner_credits')

    # Summary: clients with APC Select packages + credit balances
    today = tz.now().date()
    select_members = ClientPackage.objects.filter(
        package__package_type='select',
        status='active',
        expiry_date__gte=today,
    ).select_related('client__user', 'package').order_by('client__user__first_name')

    # All credits (paginated by most recent)
    all_credits = ClientCredit.objects.select_related(
        'client__user', 'source_package__package', 'applied_to__package', 'created_by'
    ).order_by('-created_at')[:200]

    # Available credit totals per client — DB aggregate instead of Python loop
    client_balances = dict(
        ClientCredit.objects.filter(
            status='available'
        ).filter(
            Q(expires_at__isnull=True) | Q(expires_at__gte=timezone.localdate())
        ).values('client_id').annotate(total=Sum('amount')).values_list('client_id', 'total')
    )

    clients_with_credits = Client.objects.filter(
        credits__status='available'
    ).distinct().select_related('user')

    context = {
        'select_members': select_members,
        'all_credits': all_credits,
        'client_balances': client_balances,
        'clients_with_credits': clients_with_credits,
        'all_clients': Client.objects.select_related('user').order_by('user__first_name'),
        'credit_type_choices': ClientCredit.CREDIT_TYPE_CHOICES,
    }
    return render(request, 'owner/credits.html', context)


# ============================================================================
# DISCOUNT CODES
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_discount_codes(request):
    """Manage promotional discount codes."""
    from clients.models import DiscountCode, DiscountCodeUse
    from bookings.models import SessionType

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create':
            try:
                code_str = request.POST.get('code', '').strip().upper()
                if not code_str:
                    messages.error(request, 'Code cannot be blank.')
                    return redirect('owner_discount_codes')
                dc = DiscountCode.objects.create(
                    code=code_str,
                    description=request.POST.get('description', '').strip(),
                    discount_type=request.POST.get('discount_type'),
                    value=request.POST.get('value'),
                    scope=request.POST.get('scope', 'all'),
                    max_uses=request.POST.get('max_uses') or None,
                    max_uses_per_client=int(request.POST.get('max_uses_per_client') or 1),
                    min_purchase_amount=request.POST.get('min_purchase_amount') or None,
                    valid_from=request.POST.get('valid_from') or None,
                    valid_until=request.POST.get('valid_until') or None,
                    is_active=True,
                    created_by=request.user,
                )
                pkg_ids = request.POST.getlist('specific_packages')
                if pkg_ids:
                    dc.specific_packages.set(pkg_ids)
                st_ids = request.POST.getlist('specific_session_types')
                if st_ids:
                    dc.specific_session_types.set(st_ids)
                messages.success(request, f'Discount code "{dc.code}" created.')
            except Exception as e:
                messages.error(request, f'Error creating code: {e}')

        elif action == 'edit':
            pk = request.POST.get('code_id')
            dc = get_object_or_404(DiscountCode, pk=pk)
            try:
                dc.description         = request.POST.get('description', '').strip()
                dc.discount_type       = request.POST.get('discount_type')
                dc.value               = request.POST.get('value')
                dc.scope               = request.POST.get('scope', 'all')
                dc.max_uses            = request.POST.get('max_uses') or None
                dc.max_uses_per_client = int(request.POST.get('max_uses_per_client') or 1)
                dc.min_purchase_amount = request.POST.get('min_purchase_amount') or None
                dc.valid_from          = request.POST.get('valid_from') or None
                dc.valid_until         = request.POST.get('valid_until') or None
                dc.save()
                dc.specific_packages.set(request.POST.getlist('specific_packages'))
                dc.specific_session_types.set(request.POST.getlist('specific_session_types'))
                messages.success(request, f'Code "{dc.code}" updated.')
            except Exception as e:
                messages.error(request, f'Error updating code: {e}')

        elif action == 'toggle':
            dc = get_object_or_404(DiscountCode, pk=request.POST.get('code_id'))
            dc.is_active = not dc.is_active
            dc.save(update_fields=['is_active'])
            messages.success(request, f'Code {dc.code} {"activated" if dc.is_active else "deactivated"}.')

        elif action == 'delete':
            dc = get_object_or_404(DiscountCode, pk=request.POST.get('code_id'))
            if dc.uses.filter(status='applied').exists():
                messages.error(request, f'Cannot delete "{dc.code}" — it has been used. Deactivate it instead.')
            else:
                code_str = dc.code
                dc.delete()
                messages.success(request, f'Discount code "{code_str}" deleted.')

        return redirect('owner_discount_codes')

    _AUTO_CODES = ['SIBLING-AUTO']
    codes = DiscountCode.objects.exclude(code__in=_AUTO_CODES)\
        .prefetch_related('uses', 'specific_packages', 'specific_session_types').order_by('-created_at')
    sibling_dc = DiscountCode.objects.filter(code='SIBLING-AUTO').first()
    context = {
        'codes': codes,
        'sibling_dc': sibling_dc,
        'all_packages': Package.objects.filter(is_active=True).order_by('price'),
        'all_session_types': SessionType.objects.filter(is_active=True).order_by('name'),
    }
    return render(request, 'owner/discount_codes.html', context)


@login_required
@user_passes_test(is_owner)
def owner_discount_code_detail(request, pk):
    """Usage log for a specific discount code."""
    from clients.models import DiscountCode
    dc = get_object_or_404(DiscountCode, pk=pk)
    uses = dc.uses.select_related(
        'client__user', 'applied_to_package__package', 'applied_to_booking__session_type'
    ).order_by('-used_at')
    return render(request, 'owner/discount_code_detail.html', {'code': dc, 'uses': uses})


# ============================================================================
# WAIVERS MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_waivers(request):
    """Waiver compliance dashboard — shows signed and unsigned clients."""
    today = timezone.now()
    current_year = today.year

    # Only track waivers for Client group members — exclude coaches, owners, staff
    all_clients = Client.objects.select_related('user').filter(
        user__groups__name='Client',
        user__is_staff=False,
        user__is_superuser=False,
    ).exclude(
        user__groups__name__in=['Owner', 'Coach']
    ).distinct().order_by('user__first_name')

    signed_ids = set(
        ClientWaiver.objects.filter(
            valid_year=current_year,
            waiver_version=ClientWaiver.WAIVER_VERSION,
        ).values_list('client_id', flat=True)
    )

    signed   = [c for c in all_clients if c.id in signed_ids]
    unsigned = [c for c in all_clients if c.id not in signed_ids]

    recent_waivers = ClientWaiver.objects.select_related(
        'client__user'
    ).order_by('-signed_at')[:100]

    context = {
        'signed': signed,
        'unsigned': unsigned,
        'recent_waivers': recent_waivers,
        'current_year': current_year,
        'waiver_version': ClientWaiver.WAIVER_VERSION,
        'total': all_clients.count(),
        'signed_count': len(signed),
    }
    return render(request, 'owner/waivers.html', context)


# ============================================================================
# CONTACT IMPORT MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_contacts(request):
    """Imported contact registry — parents from past events with their players."""
    from clients.models import ContactParent, ContactPlayer
    from django.db.models import Count, Prefetch

    search  = request.GET.get('q', '').strip()
    status  = request.GET.get('status', '')   # linked / unlinked
    source  = request.GET.get('source', '')

    qs = ContactParent.objects.prefetch_related('players').annotate(
        annotated_player_count=Count('players', distinct=True)
    ).order_by('last_name', 'first_name', 'email')

    if search:
        from django.db.models import Q
        qs = qs.filter(
            Q(email__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(phone__icontains=search) |
            Q(players__first_name__icontains=search) |
            Q(players__last_name__icontains=search)
        ).distinct()

    if status == 'linked':
        qs = qs.filter(client__isnull=False)
    elif status == 'unlinked':
        qs = qs.filter(client__isnull=True)

    if source:
        qs = qs.filter(source=source)

    total         = ContactParent.objects.count()
    linked_count  = ContactParent.objects.filter(client__isnull=False).count()
    player_count  = ContactPlayer.objects.count()

    context = {
        'contacts':      qs,
        'total':         total,
        'linked_count':  linked_count,
        'unlinked_count':total - linked_count,
        'player_count':  player_count,
        'search':        search,
        'status':        status,
        'source_filter': source,
        'source_choices':ContactParent.SOURCE_CHOICES,
    }
    return render(request, 'owner/contacts.html', context)


@login_required
@user_passes_test(is_owner)
def owner_contact_edit(request, pk):
    """Edit a ContactParent and all their associated players."""
    from clients.models import ContactParent, ContactPlayer
    contact = get_object_or_404(ContactParent, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action', 'save_parent')

        if action == 'save_parent':
            contact.first_name = request.POST.get('first_name', '').strip()
            contact.last_name  = request.POST.get('last_name', '').strip()
            contact.email      = request.POST.get('email', '').strip()
            contact.phone      = request.POST.get('phone', '').strip()
            contact.source     = request.POST.get('source', contact.source)
            contact.notes      = request.POST.get('notes', '').strip()
            contact.save()
            messages.success(request, 'Contact updated.')

        elif action == 'save_player':
            player_id = request.POST.get('player_id')
            if player_id:
                player = get_object_or_404(ContactPlayer, pk=player_id, parent=contact)
            else:
                player = ContactPlayer(parent=contact)
            player.first_name  = request.POST.get('first_name', '').strip()
            player.last_name   = request.POST.get('last_name', '').strip()
            by = request.POST.get('birth_year', '').strip()
            player.birth_year  = int(by) if by.isdigit() else None
            player.sex         = request.POST.get('sex', '')
            player.club_team   = request.POST.get('club_team', '').strip()
            player.position    = request.POST.get('position', '').strip()
            player.notes       = request.POST.get('notes', '').strip()
            player.save()
            messages.success(request, f'Player {"added" if not player_id else "updated"}.')

        elif action == 'delete_player':
            player_id = request.POST.get('player_id')
            ContactPlayer.objects.filter(pk=player_id, parent=contact).delete()
            messages.success(request, 'Player removed.')

        elif action == 'delete_contact':
            contact.delete()
            messages.success(request, 'Contact deleted.')
            return redirect('owner_contacts')

        return redirect('owner_contact_edit', pk=contact.pk)

    context = {
        'contact':  contact,
        'players':  contact.players.order_by('last_name', 'first_name'),
        'source_choices': ContactParent.SOURCE_CHOICES,
        'sex_choices': ContactPlayer.SEX_CHOICES,
    }
    return render(request, 'owner/contact_edit.html', context)


# ============================================================================
# OWNER GUIDE
# ============================================================================

@login_required
@user_passes_test(is_owner)
def owner_guide(request):
    """Owner how-to guide."""
    return render(request, 'owner/guide.html')
