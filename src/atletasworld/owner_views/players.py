from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db.models import Count, Q
from clients.models import Player, ClientPackage
from bookings.models import Booking
from coaches.models import PlayerAssessment
from ._auth import is_owner


@login_required
@user_passes_test(is_owner)
def owner_players(request):
    """List all players with session-type / package filters and CSV/XLSX/PDF export."""
    from clients.models import ClientPackage
    from bookings.models import Booking, SessionType as ST

    # ── Build base queryset ──────────────────────────────────────────────────
    qs = Player.objects.select_related('client__user').annotate(
        total_bookings=Count('bookings', distinct=True),
        total_assessments=Count('assessments', distinct=True),
    )

    # ── Filters ──────────────────────────────────────────────────────────────
    session_type_id = request.GET.get('session_type')
    package_id      = request.GET.get('package')
    search          = request.GET.get('q', '').strip()

    if session_type_id:
        qs = qs.filter(bookings__session_type_id=session_type_id).distinct()
    if package_id:
        today = timezone.localdate()
        qs = qs.filter(
            client__packages__package_id=package_id,
            client__packages__status='active',
            client__packages__expiry_date__gte=today,
        ).distinct()
    if search:
        qs = qs.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(client__user__email__icontains=search)
        )

    players = qs.order_by('first_name', 'last_name')

    # ── Export ───────────────────────────────────────────────────────────────
    export = request.GET.get('export')
    if export in ('csv', 'xlsx', 'pdf'):
        # Load full data for export (no limit)
        rows = []
        for p in players.prefetch_related('client__packages__package'):
            active_pkgs = ', '.join(
                cp.package.name
                for cp in p.client.packages.filter(status='active').select_related('package')
            )
            rows.append({
                'First Name':      p.first_name,
                'Last Name':       p.last_name,
                'Birth Year':      p.birth_year,
                'Age Group':       p.age_group,
                'Gender':          p.get_gender_display(),
                'Position':        p.get_primary_position_display() or '',
                'Skill Level':     p.get_skill_level_display(),
                'Soccer Club':     p.soccer_club,
                'Team Name':       p.team_name,
                'Jersey Size':     p.get_jersey_size_display() if p.jersey_size else '',
                'National Team':   p.favorite_national_team,
                'Club Team':       p.favorite_club_team,
                'Parent/Guardian': p.client.user.get_full_name() or p.client.user.username,
                'Parent Email':    p.client.user.email,
                'Parent Phone':    p.client.phone,
                'Total Bookings':  p.total_bookings,
                'Total Assessments': p.total_assessments,
                'Active Packages': active_pkgs,
                'Notes':           p.notes,
            })

        headers = list(rows[0].keys()) if rows else []

        if export == 'csv':
            import csv
            from django.http import HttpResponse
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="players.csv"'
            writer = csv.DictWriter(response, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
            return response

        elif export == 'xlsx':
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse

            wb = Workbook()
            ws = wb.active
            ws.title = 'Players'

            # Header row — styled
            header_fill = PatternFill('solid', fgColor='6366F1')
            header_font = Font(bold=True, color='FFFFFF')
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(h) + 4)

            # Data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row[h])

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="players.xlsx"'
            return response

        elif export == 'pdf':
            # Render a print-optimized HTML page — browser handles PDF conversion
            context = {'players': players, 'rows': rows, 'headers': headers,
                       'total': players.count()}
            return render(request, 'owner/players_print.html', context)

    # ── Normal page view (limit to 500 for performance) ──────────────────────
    filter_session_types = ST.objects.filter(is_active=True).order_by('name')
    from clients.models import Package
    filter_packages = Package.objects.filter(is_active=True).order_by('name')

    context = {
        'players':             players[:500],
        'total':               players.count(),
        'filter_session_types': filter_session_types,
        'filter_packages':     filter_packages,
        'selected_session':    session_type_id or '',
        'selected_package':    package_id or '',
        'search':              search,
    }
    return render(request, 'owner/players.html', context)


@login_required
@user_passes_test(is_owner)
def owner_player_detail(request, pk):
    """View a single player's profile, bookings, packages, and assessments."""
    player = get_object_or_404(Player.objects.select_related('client__user', 'team'), pk=pk)
    bookings = Booking.objects.filter(player=player).select_related('coach__user', 'session_type').order_by('-scheduled_date')[:30]
    from clients.models import ClientPackage
    packages = ClientPackage.objects.filter(client=player.client).select_related('package').order_by('-purchase_date')
    assessments = PlayerAssessment.objects.filter(player=player).select_related('coach__user').order_by('-assessment_date')[:10]
    context = {
        'player': player,
        'bookings': bookings,
        'packages': packages,
        'assessments': assessments,
    }
    return render(request, 'owner/player_detail.html', context)
